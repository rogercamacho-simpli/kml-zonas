import streamlit as st
import openpyxl
import io
import re
import requests
import json
import math
from datetime import date, datetime

st.set_page_config(page_title="SimpliRoute Tools", page_icon="🚀", layout="wide")

# ── NAVEGACIÓN ────────────────────────────────────────────────────────────────
MENU_GROUPS = [
    ("📋 Visitas", [
        "🧑‍💼 Agregar Seller a Visitas",
        "✏️ Edición de Visitas",
        "🗑️ Eliminación Masiva de Visitas",
        "🏷️ Tipos de Visita y Skills",
    ]),
    ("🚗 Rutas & Flotas", [
        "🚛 Flotas",
        "🚦 Iniciar / Cerrar Rutas",
        "🗺️ Zonas",
    ]),
    ("📡 GPS", [
        "📍 Análisis de Recorrido GPS",
        "📡 Validación de GPS",
    ]),
    ("⚙️ Configuración", [
        "👤 Cambiar Rol de Usuario",
        "⚙️ Configurar Addons",
        "🔔 Crear Webhook",
        "🔓 Desbloqueo de Contraseña",
        "🔍 Permisos de Usuario",
        "🔁 Reenviar Webhooks",
    ]),
    ("📦 TMS", [
        "📄 Tipos de Documento",
        "🚚 Transportistas",
    ]),
]

if "current_page" not in st.session_state:
    st.session_state["current_page"] = MENU_GROUPS[0][1][0]

selected = st.session_state["current_page"]

def nav_item(label):
    is_active = st.session_state["current_page"] == label
    if st.button(label, key=f"nav_{label}", use_container_width=True,
                 type="primary" if is_active else "secondary"):
        for key in list(st.session_state.keys()):
            if key != "current_page":
                del st.session_state[key]
        st.session_state["current_page"] = label
        st.rerun()

with st.sidebar:
    st.markdown("## 🚀 SimpliRoute Tools")
    st.markdown("---")
    for group_name, items in MENU_GROUPS:
        is_group_active = selected in items
        with st.expander(group_name, expanded=is_group_active):
            for item in items:
                nav_item(item)
    st.markdown("---")
    st.link_button("🔧 Tools Julio", "https://simpliroute-tools.streamlit.app/", use_container_width=True)
    st.markdown("---")
    st.caption("SimpliRoute Internal Tools v1.0")


# ── HELPERS ───────────────────────────────────────────────────────────────────
def haversine_m(lat1, lon1, lat2, lon2):
    R = 6371000
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lon2 - lon1)
    a = math.sin(dphi/2)**2 + math.cos(phi1)*math.cos(phi2)*math.sin(dlambda/2)**2
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))

def decode_file(raw_bytes):
    for enc in ["utf-8", "latin-1", "cp1252"]:
        try:
            return raw_bytes.decode(enc, errors="ignore")
        except Exception:
            continue
    return raw_bytes.decode("utf-8", errors="ignore")

def strip_rtf_codes(text):
    text = re.sub(r"\\'[0-9a-fA-F]{2}", "", text)
    text = re.sub(r"\\[a-zA-Z]+\d*\s?", " ", text)
    text = re.sub(r"[{}]", "", text)
    text = re.sub(r" {2,}", " ", text)
    return text

def parse_polygons(text):
    polygons = []
    for block in re.findall(r'<Placemark>(.*?)</Placemark>', text, re.DOTALL):
        if '<Polygon>' not in block and '<Polygon ' not in block:
            continue
        name_match = re.search(r'<name>\s*(.*?)\s*</name>', block)
        name = name_match.group(1).strip() if name_match else "Sin nombre"
        coords_match = re.search(r'<coordinates>\s*(.*?)\s*</coordinates>', block, re.DOTALL)
        if not coords_match:
            continue
        points = []
        for token_str in coords_match.group(1).strip().split():
            parts = token_str.strip().split(",")
            if len(parts) < 2:
                continue
            try:
                a, b = float(parts[0]), float(parts[1])
            except ValueError:
                continue
            lng, lat = a, b
            if abs(lat) > 90:
                lng, lat = lat, lng
            if -90 <= lat <= 90 and -180 <= lng <= 180:
                points.append((lat, lng))
        if points:
            polygons.append({"name": name, "coords": points})
    return polygons

def coords_to_str(coords):
    return "[" + ",".join(f"{{'lat':'{lat:.6f}','lng':'{lng:.6f}'}}" for lat, lng in coords) + "]"

def generate_excel_zones(polygons):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Hoja 1"
    ws.append(["is_name", "is_coordinates"])
    for p in polygons:
        ws.append([p["name"], coords_to_str(p["coords"])])
    ws.column_dimensions["A"].width = 25; ws.column_dimensions["B"].width = 80
    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf

def read_excel_column(file, columns):
    wb = openpyxl.load_workbook(file); ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        entry = {}
        for col in columns:
            idx = headers.index(col) if col in headers else -1
            val = row[idx] if 0 <= idx < len(row) else None
            entry[col] = str(val).strip() if val is not None else ""
        if any(entry[c] for c in columns):
            rows.append(entry)
    return rows

def show_results(results, name_key):
    ok = sum(1 for r in results if r["code"] in [200, 201])
    st.markdown(f"✅ **{ok} creados** | ❌ **{len(results)-ok} con error**")
    for r in results:
        if r["code"] in [200, 201]: st.success(f"✅ {r[name_key]} — OK")
        elif r["code"] == 400: st.warning(f"⚠️ {r[name_key]} — Ya existe o datos inválidos")
        elif r["code"] == 401: st.error(f"❌ {r[name_key]} — Token inválido")
        elif r["code"] is None: st.error(f"❌ {r[name_key]} — Sin conexión")
        else: st.error(f"❌ {r[name_key]} — Error {r['code']}")

def get_users_list(token):
    try:
        r = requests.get("http://api.simpliroute.com/v1/accounts/users/",
                         headers={"Authorization": f"Token {token}", "accept": "application/json"}, timeout=300)
        return r.status_code, r.json()
    except Exception as e:
        return None, str(e)

def put_user_full(token, user, new_email):
    try:
        r = requests.put(f"http://api.simpliroute.com/v1/accounts/users/{user['id']}/",
                         headers={"Authorization": f"Token {token}", "Content-Type": "application/json"},
                         json={"id": user["id"], "username": user.get("username",""), "name": user.get("name",""),
                               "phone": user.get("phone",""), "email": new_email,
                               "is_owner": user.get("is_owner",False), "is_admin": user.get("is_admin",False),
                               "is_driver": user.get("is_driver",False), "is_codriver": user.get("is_codriver",False),
                               "is_router_jr": user.get("is_router_jr",False), "is_monitor": user.get("is_monitor",False),
                               "is_coordinator": user.get("is_coordinator",False), "is_router": user.get("is_router",False),
                               "is_staff": user.get("is_staff",False), "is_seller_viewer": user.get("is_seller_viewer",False),
                               "is_seller": user.get("is_seller",False), "blocked": user.get("blocked",False),
                               "status": user.get("status","active")}, timeout=30)
        return r.status_code, r.json()
    except Exception as e:
        return None, str(e)


# ── MASTER LIST DE ADDONS ─────────────────────────────────────────────────────
MASTER_ADDONS = [
    {"key": "visit_signature",            "title": "Firma de Visita",              "description": "",                             "logo": ""},
    {"key": "path_finder",                "title": "path_finder",                  "description": "path_finder",                  "logo": ""},
    {"key": "ada",                        "title": "Ada",                          "description": "ada",                          "logo": "ada"},
    {"key": "live_tracking",              "title": "Live Tracking",                "description": "Live Tracking",                "logo": ""},
    {"key": "rest_time",                  "title": "Tiempo de Descanso",           "description": "rest_time",                    "logo": ""},
    {"key": "mobile_route_edition",       "title": "Edición de Ruta Móvil",        "description": "Addon mobile_route_edition",   "logo": "-"},
    {"key": "mobile_features",            "title": "Funciones Móviles",            "description": "Addon mobile_features",        "logo": "-"},
    {"key": "seller_management",          "title": "Gestión de Sellers",           "description": "seller_management",            "logo": ""},
    {"key": "territory_planner",          "title": "force field",                  "description": "force field",                  "logo": ""},
    {"key": "account_configs_edition",    "title": "Edición de Configuración",     "description": "account_configs_edition",      "logo": ""},
    {"key": "management_comments",        "title": "Comentarios",                  "description": "comentarios",                  "logo": ""},
    {"key": "accounting",                 "title": "accounting",                   "description": "accounting",                   "logo": ""},
    {"key": "simpli_chat",                "title": "Simpli Chat",                  "description": "SImpli Chat",                  "logo": ""},
    {"key": "zones",                      "title": "Zonas",                        "description": "Addon zones",                  "logo": ""},
    {"key": "delivery_survey",            "title": "Delivery Survey",              "description": "delivery_survey",              "logo": ""},
    {"key": "tms",                        "title": "TMS",                          "description": "tms",                          "logo": ""},
    {"key": "visit_card_additional_data", "title": "Visit Card Additional Data",   "description": "visit_card_additional_data",   "logo": "-"},
    {"key": "roles",                      "title": "Roles Personalizados",         "description": "roles",                        "logo": ""},
]


# ── FEATURE: CONFIGURAR ADDONS ────────────────────────────────────────────────
def page_configurar_addons():
    st.title("⚙️ Configurar Addons")
    token = st.text_input("🔑 Token de SimpliRoute", type="password", key="token_addons")

    if st.button("🔍 Consultar Addons", type="primary", disabled=not token):
        with st.spinner("Consultando addons..."):
            try:
                r = requests.get("http://api.simpliroute.com/v1/addons/addons/",
                                 headers={"Authorization": f"Token {token}", "accept": "application/json"}, timeout=30)
                code, resp = r.status_code, r.json()
            except Exception as e:
                code, resp = None, str(e)
        if code == 200:
            acc_id = resp[0]["account_id"] if resp else "—"
            st.session_state["addons_data"]    = resp
            st.session_state["addons_token"]   = token
            st.session_state["addons_account"] = acc_id
        elif code == 401:
            st.error("❌ Token inválido")
        else:
            st.error(f"❌ Error {code}: {resp}")
            st.session_state.pop("addons_data", None)

    if "addons_data" in st.session_state:
        addons        = st.session_state["addons_data"]
        tok           = st.session_state["addons_token"]
        acc_id        = st.session_state["addons_account"]
        existing_keys = {a["key"] for a in addons}

        st.info(f"🏢 Cuenta: **{acc_id}** · {len(addons)} addon(s) configurados")
        st.divider()
        st.subheader("📋 Addons de la cuenta")

        cols = st.columns(3)
        for idx, addon in enumerate(addons):
            master_entry = next((m for m in MASTER_ADDONS if m["key"] == addon["key"]), None)
            label  = master_entry["title"] if master_entry else (addon.get("title") or addon.get("key"))
            estado = addon.get("enable", False)

            color_borde  = "#1D9E75" if estado else "#E24B4A"
            status_color = "#0F6E56" if estado else "#A32D2D"
            badge_bg     = "#E1F5EE" if estado else "#FCEBEB"
            estado_txt   = "✅ Activo" if estado else "🔴 Inactivo"
            btn_label    = "🔴 Desactivar" if estado else "🟢 Activar"

            with cols[idx % 3]:
                st.markdown(
                    f"""<div style="border:1.5px solid {color_borde};border-radius:10px;padding:14px 16px 8px 16px;margin-bottom:12px;background:#ffffff;">
                        <div style="font-size:14px;font-weight:600;color:#1a1a1a;margin-bottom:4px;">{label}</div>
                        <div style="font-size:11px;font-family:monospace;background:{badge_bg};color:{status_color};display:inline-block;padding:1px 7px;border-radius:4px;margin-bottom:8px;">{addon['key']}</div>
                        <div style="font-size:13px;color:{status_color};font-weight:500;">{estado_txt}</div>
                    </div>""", unsafe_allow_html=True)
                if st.button(btn_label, key=f"toggle_{addon['id']}", use_container_width=True):
                    new_val = not estado
                    payload = [{"id": addon["id"], "account_id": addon["account_id"], "title": addon.get("title",""),
                                "key": addon["key"], "description": addon.get("description",""),
                                "logo": addon.get("logo",""), "value": new_val}]
                    try:
                        r = requests.put("http://api.simpliroute.com/v1/addons/addons/",
                                         headers={"Authorization": f"Token {tok}", "Content-Type": "application/json"},
                                         json=payload, timeout=15)
                        put_code = r.status_code
                    except:
                        put_code = None
                    if put_code in [200, 201]:
                        for a in st.session_state["addons_data"]:
                            if a["id"] == addon["id"]: a["enable"] = new_val
                        st.rerun()
                    elif put_code == 401: st.error("❌ Token inválido")
                    else: st.error(f"❌ Error {put_code}")

        missing = [m for m in MASTER_ADDONS if m["key"] not in existing_keys]
        if missing:
            st.divider()
            st.subheader("➕ Addons no configurados")
            st.caption("Estos addons no están en la cuenta. Puedes agregarlos aquí.")
            staff_tok_input = st.text_input("🔐 Token Staff *(requerido para agregar addons)*", type="password",
                                            key="staff_token_addons", placeholder="Token de usuario Staff")
            cols2 = st.columns(3)
            for idx2, master in enumerate(missing):
                with cols2[idx2 % 3]:
                    st.markdown(
                        f"""<div style="border:1.5px dashed #ccc;border-radius:10px;padding:14px 16px 8px 16px;margin-bottom:12px;background:#fafafa;">
                            <div style="font-size:14px;font-weight:600;color:#555;margin-bottom:4px;">{master['title']}</div>
                            <div style="font-size:11px;font-family:monospace;background:#f0f0f0;color:#888;display:inline-block;padding:1px 7px;border-radius:4px;margin-bottom:8px;">{master['key']}</div>
                            <div style="font-size:12px;color:#aaa;">No configurado</div>
                        </div>""", unsafe_allow_html=True)
                    if st.button("➕ Agregar", key=f"add_{master['key']}", use_container_width=True):
                        use_tok = staff_tok_input.strip() if staff_tok_input and staff_tok_input.strip() else tok
                        payload = {"account_id": int(acc_id), "title": master["title"], "key": master["key"],
                                   "description": master["description"], "logo": master["logo"], "value": True}
                        try:
                            r = requests.post("http://api.simpliroute.com/v1/addons/addons/",
                                              headers={"Authorization": f"Token {use_tok}", "Content-Type": "application/json"},
                                              json=payload, timeout=15)
                            post_code, post_resp = r.status_code, r.json()
                        except Exception as e:
                            post_code, post_resp = None, str(e)
                        if post_code in [200, 201]:
                            st.success(f"✅ **{master['title']}** agregado")
                            try:
                                r2 = requests.get("http://api.simpliroute.com/v1/addons/addons/",
                                                  headers={"Authorization": f"Token {tok}", "accept": "application/json"}, timeout=15)
                                if r2.status_code == 200: st.session_state["addons_data"] = r2.json()
                            except: pass
                            st.rerun()
                        elif post_code == 403:
                            st.error("❌ Error 403 — Esta acción solo puede realizarla un usuario Staff. Ingresa tu token de Staff en el campo de arriba.")
                        elif post_code == 401: st.error("❌ Token inválido")
                        else: st.error(f"❌ Error {post_code}: {post_resp}")
        else:
            st.divider()
            st.info("✅ La cuenta tiene todos los addons del catálogo configurados.")


# ── FEATURE: PERMISOS DE USUARIO ─────────────────────────────────────────────
PERMISOS_AMIGABLES = {
    "reports.edit":                               "📊 Reportes",
    "watchtower.tracking.edit":                   "🗺️ Watch Tower - Tracking",
    "watchtower.visits.edit":                     "📍 Watch Tower - Visitas",
    "settings.communications.live_tracking.edit": "📡 Comunicaciones",
    "settings.communications.wsp_pro.edit":       "💬 WhatsApp Pro",
    "clients.edit":                               "👥 Clientes",
    "chat.edit":                                  "💬 Chat",
    "ada.edit":                                   "🤖 Ada",
    "forcefield.edit":                            "🛡️ ForceField",
    "settings.vehicles.edit":                     "🚗 Vehículos",
    "pathfinder.plans.edit":                      "📋 Mis planes",
    "pathfinder.router.edit":                     "🔀 Ruteador",
    "pathfinder.assigner.edit":                   "📌 Asignador",
    "settings.users.edit":                        "👤 Usuarios",
    "settings.vehicles.fleets.edit":              "🚛 Flotas",
    "settings.zones.edit":                        "🗺️ Zonas",
    "settings.optimizers.edit":                   "⚙️ Creación de rutas",
    "settings.mobile.edit":                       "📱 App móvil",
    "settings.security.edit":                     "🔒 Seguridad",
    "settings.customization.edit":                "🎨 Personalización",
}

SYSTEM_ROLE_FLAGS = {
    "is_owner":         "Owner",
    "is_admin":         "Administrador",
    "is_router":        "Router",
    "is_router_jr":     "Router Jr",
    "is_coordinator":   "Coordinador",
    "is_monitor":       "Monitor",
    "is_seller":        "Seller",
    "is_seller_viewer": "Seller Viewer",
    "is_driver":        "Conductor",
    "is_codriver":      "Co-conductor",
}

def page_permisos_usuario():
    st.title("🔍 Permisos de Usuario")
    st.markdown("Consulta el rol y permisos de un usuario, incluyendo roles personalizados.")

    col1, col2 = st.columns(2)
    with col1: token   = st.text_input("🔑 Token", type="password", key="token_permisos")
    with col2: user_id = st.text_input("🆔 User ID", placeholder="Ej: 503734", key="user_permisos")

    if not st.button("🔍 Consultar", type="primary", disabled=not (token and user_id)):
        return

    # 1. Obtener account_id via /me
    try:
        rme = requests.get("http://api.simpliroute.com/v1/accounts/me",
                           headers={"Authorization": f"Token {token}", "Content-Type": "application/json"}, timeout=15)
        if rme.status_code == 401:
            st.error("❌ Token inválido"); return
        if rme.status_code != 200:
            st.error(f"❌ Error obteniendo cuenta: {rme.status_code}"); return
        account_id = rme.json().get("account", {}).get("id")
        if not account_id:
            st.error("❌ No se pudo obtener el Account ID"); return
    except Exception as e:
        st.error(f"❌ Error: {e}"); return

    # 2. Obtener usuario
    try:
        ru = requests.get(f"http://api.simpliroute.com/v1/accounts/users/{user_id.strip()}/",
                          headers={"Authorization": f"Token {token}"}, timeout=15)
        if ru.status_code == 401:
            st.error("❌ Token inválido"); return
        if ru.status_code != 200:
            st.error(f"❌ Error {ru.status_code}"); return
        user = ru.json()
    except Exception as e:
        st.error(f"❌ Error consultando usuario: {e}"); return

    # Info del usuario
    st.divider()
    c1, c2, c3 = st.columns(3)
    c1.markdown(f"**Nombre**<br>{user.get('name','—')}", unsafe_allow_html=True)
    c2.markdown(f"**Username**<br>{user.get('username','—')}", unsafe_allow_html=True)
    estado = "🔴 Bloqueado" if user.get("blocked") else "🟢 Activo"
    c3.markdown(f"**Estado**<br>{estado}", unsafe_allow_html=True)
    st.caption(f"📧 {user.get('email') or 'Sin email'} · ID: {user.get('id')} · Cuenta: {account_id}")

    # 3. Obtener roles de la cuenta
    roles = []
    try:
        rr = requests.get(f"https://api-gateway.simpliroute.com/v1/accounts/{account_id}/roles/",
                          headers={"Authorization": f"Token {token}", "accept": "application/json"}, timeout=15)
        if rr.status_code == 200:
            roles = rr.json()
    except Exception:
        pass

    # 4. Determinar rol y permisos
    custom_role_name = user.get("custom_role_name")
    st.divider()

    if custom_role_name:
        role_data = next((r for r in roles if r.get("name") == custom_role_name), None)
        st.markdown(f"**🏷️ Rol:** `{custom_role_name}` — *personalizado*")
    else:
        flag_key   = next((f for f in SYSTEM_ROLE_FLAGS if user.get(f)), None)
        role_label = SYSTEM_ROLE_FLAGS.get(flag_key, "Sin rol")
        role_data  = next((r for r in roles if r.get("legacy_role") == flag_key and r.get("is_system_role")), None)
        st.markdown(f"**🏷️ Rol:** `{role_label}` — *sistema*")

    if role_data:
        perm_keys = [p["key"] for p in role_data.get("permissions", [])]
        tiene    = [PERMISOS_AMIGABLES[k] for k in PERMISOS_AMIGABLES if k in perm_keys]
        no_tiene = [PERMISOS_AMIGABLES[k] for k in PERMISOS_AMIGABLES if k not in perm_keys]

        st.markdown(f"**✅ Tiene acceso a ({len(tiene)}):**")
        cols = st.columns(2)
        for i, label in enumerate(tiene):
            cols[i % 2].markdown(f"• {label}")

        if no_tiene:
            with st.expander(f"🚫 Sin acceso ({len(no_tiene)})"):
                cols2 = st.columns(2)
                for i, label in enumerate(no_tiene):
                    cols2[i % 2].markdown(f"• {label}")
    else:
        st.warning("⚠️ No se encontraron los permisos detallados para este rol.")


# ── FEATURE: AGREGAR SELLER A VISITAS ────────────────────────────────────────
def page_agregar_seller():
    st.title("🧑‍💼 Agregar Seller a Visitas")
    token = st.text_input("🔑 Token de SimpliRoute", type="password", key="token_seller")
    if st.button("🔍 Consultar Sellers", type="primary", disabled=not token):
        with st.spinner("Consultando..."):
            try:
                r = requests.get("http://api.simpliroute.com/v1/sellers/",
                                 headers={"Authorization": f"Token {token}"}, timeout=60)
                code, resp = r.status_code, r.json()
            except Exception as e:
                code, resp = None, str(e)
        if code == 200:
            if not resp:
                st.warning("⚠️ Esta cuenta no tiene sellers configurados.")
                st.session_state.pop("sellers", None)
            else:
                st.session_state["sellers"] = resp; st.session_state["seller_token"] = token
                st.success(f"✅ {len(resp)} seller(s)")
        else:
            st.error(f"❌ Error {code}"); st.session_state.pop("sellers", None)

    if "sellers" in st.session_state:
        sellers = st.session_state["sellers"]
        st.divider()
        def slabel(s):
            name = s.get("name") or s.get("username") or s.get("email") or str(s.get("uuid",""))
            email = s.get("email","")
            return f"{name} — {email}" if email and email != name else name
        opciones = {slabel(s): s for s in sellers}
        selected_s = opciones[st.selectbox("Seller:", list(opciones.keys()))]
        seller_uuid = selected_s.get("uuid") or selected_s.get("id")
        st.divider()
        st.caption("Pega los IDs de visita uno por línea:\n```\n799841373\n808472905\n```")
        visit_ids_raw = st.text_area("IDs", placeholder="799841373\n808472905", height=200, label_visibility="collapsed")
        if st.button("💾 Asignar Seller", type="primary", disabled=not visit_ids_raw):
            ids, errs = [], []
            for l in visit_ids_raw.strip().splitlines():
                l = l.strip()
                try: ids.append(int(l))
                except: errs.append(l)
            if errs: st.error(f"❌ IDs inválidos: {', '.join(errs)}"); return
            with st.spinner(f"Asignando a {len(ids)} visitas..."):
                try:
                    r = requests.patch("http://api.simpliroute.com/v1/routes/visits/",
                                       headers={"Authorization": f"Token {st.session_state['seller_token']}", "Content-Type": "application/json"},
                                       json=[{"id": vid, "seller": seller_uuid} for vid in ids], timeout=120)
                    code = r.status_code
                except: code = None
            if code in [200,201]: st.success(f"✅ Seller asignado a {len(ids)} visitas")
            else: st.error(f"❌ Error {code}")


# ── FEATURE: FLOTAS ───────────────────────────────────────────────────────────
def page_flotas():
    st.title("🚛 Flotas")
    tab1, tab2 = st.tabs(["📋 Asignar Flotas", "🗑️ Eliminar Flotas"])

    with tab1:
        st.info("ℹ️ Cada fila representa una edición independiente. Si la misma flota aparece dos veces, prevalece la última.")
        token = st.text_input("🔑 Token de SimpliRoute", type="password", key="token_flotas")
        def make_tpl():
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Flotas"
            ws.append(["Nombre de flota","Vehículos","Usuarios"])
            ws.append(["Flota Norte","MC4327,QC4380","juan.perez,ana.lopez"])
            ws.column_dimensions["A"].width=25; ws.column_dimensions["B"].width=40; ws.column_dimensions["C"].width=40
            buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf
        st.download_button("📥 Descargar plantilla", data=make_tpl(), file_name="plantilla_flotas.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        fleet_file = st.file_uploader("📂 Sube tu Excel", type=["xlsx"], key="upload_flotas")
        if fleet_file and token:
            if st.button("🚀 Procesar flotas", type="primary"):
                try:
                    rows = read_excel_column(fleet_file, ["Nombre de flota","Vehículos","Usuarios"])
                except Exception as e:
                    st.error(f"❌ Error leyendo Excel: {e}"); return
                with st.spinner("⏳ Consultando flotas, vehículos y usuarios..."):
                    try:
                        rf = requests.get("http://api.simpliroute.com/v1/fleets/", headers={"Authorization":f"Token {token}"}, timeout=300)
                        rv = requests.get("http://api.simpliroute.com/v1/routes/vehicles/", headers={"Authorization":f"Token {token}"}, timeout=300)
                        code_u, users = get_users_list(token)
                        fleets, vehicles = rf.json(), rv.json()
                    except Exception as e:
                        st.error(f"❌ Error: {e}"); return
                fleet_map = {f["name"].strip().lower(): f for f in fleets}
                vehicle_map = {v["name"].strip().lower(): v["id"] for v in vehicles if v.get("name")}
                user_map = {}
                for u in (users if isinstance(users, list) else []):
                    if u.get("username"): user_map[u["username"].strip().lower()] = u["id"]
                    if u.get("email"): user_map[u["email"].strip().lower()] = u["id"]
                st.success(f"✅ {len(fleets)} flotas · {len(vehicles)} vehículos · {len(user_map)} usuarios")
                st.divider()
                for row in rows:
                    fname = row["Nombre de flota"].strip()
                    if fname.lower() not in fleet_map:
                        st.error(f"❌ **{fname}** — Flota no encontrada"); continue
                    fleet = fleet_map[fname.lower()]
                    vids, verrs = [], []
                    for v in [x.strip() for x in row["Vehículos"].split(",") if x.strip()]:
                        vid = vehicle_map.get(v.lower())
                        (vids if vid else verrs).append(vid if vid else v)
                    uids, uerrs = [], []
                    for u in [x.strip() for x in row["Usuarios"].split(",") if x.strip()]:
                        uid = user_map.get(u.lower())
                        (uids if uid else uerrs).append(uid if uid else u)
                    if verrs: st.warning(f"⚠️ **{fname}** — Vehículos no encontrados: {', '.join(str(x) for x in verrs)}")
                    if uerrs: st.warning(f"⚠️ **{fname}** — Usuarios no encontrados: {', '.join(str(x) for x in uerrs)}")
                    try:
                        r = requests.put(f"http://api.simpliroute.com/v1/fleets/{fleet['id']}/",
                                         headers={"Authorization":f"Token {token}","Content-Type":"application/json"},
                                         json={"id":fleet["id"],"name":fname,"vehicles":vids,"users":uids}, timeout=300)
                        code = r.status_code
                    except: code = None
                    if code == 200: st.success(f"✅ **{fname}** — Actualizada ({len(vids)} vehículos · {len(uids)} usuarios)")
                    else: st.error(f"❌ **{fname}** — Error {code}")
        elif fleet_file and not token:
            st.warning("⚠️ Ingresa tu token.")

    with tab2:
        st.error("⚠️ **ADVERTENCIA:** La eliminación de flotas es permanente y no se puede deshacer.")
        token_del = st.text_input("🔑 Token de SimpliRoute", type="password", key="token_flotas_del")
        st.caption("Pega los IDs de flota uno por línea:\n```\n47524\n47525\n```")
        fleet_ids_raw = st.text_area("IDs de flota", placeholder="47524\n47525", height=180,
                                     label_visibility="collapsed", key="fleet_ids_del")
        confirm_del = st.checkbox("✅ Confirmo que quiero eliminar estas flotas de forma permanente", key="confirm_flotas_del")
        if st.button("🗑️ Eliminar Flotas", type="primary", disabled=not (token_del and fleet_ids_raw and confirm_del)):
            fleet_ids = []
            for l in fleet_ids_raw.strip().splitlines():
                l = l.strip()
                if not l: continue
                try: fleet_ids.append(int(l))
                except: st.warning(f"⚠️ ID inválido ignorado: {l}")
            if not fleet_ids: st.error("❌ No se encontraron IDs válidos."); return
            prog = st.progress(0); status = st.empty()
            ok_count = 0; err_count = 0
            for i, fid in enumerate(fleet_ids):
                status.info(f"Eliminando flota **{fid}** ({i+1}/{len(fleet_ids)})...")
                try:
                    r = requests.delete(f"https://api.simpliroute.com/v1/fleets/{fid}",
                                        headers={"Authorization": f"Token {token_del}", "accept": "application/json"}, timeout=30)
                    code = r.status_code
                except Exception as e:
                    st.error(f"❌ **{fid}** — Error: {e}"); err_count += 1; prog.progress((i+1)/len(fleet_ids)); continue
                if code in [200, 204]: st.success(f"✅ Flota **{fid}** eliminada"); ok_count += 1
                elif code == 401: st.error("❌ Token inválido."); status.empty(); prog.empty(); return
                elif code == 404: st.warning(f"⚠️ **{fid}** — No encontrada")
                else: st.error(f"❌ **{fid}** — Error {code}"); err_count += 1
                prog.progress((i+1)/len(fleet_ids))
            status.empty(); prog.empty()
            st.divider()
            if err_count == 0: st.success(f"✅ Completado — **{ok_count} flotas eliminadas**")
            else: st.warning(f"⚠️ Completado con errores — **{ok_count} eliminadas**, {err_count} con error")


# ── FEATURE: ZONAS ────────────────────────────────────────────────────────────
def page_zonas():
    st.title("🗺️ Zonas")
    tab1, tab2, tab3 = st.tabs(["📂 Cargar Zonas", "🗑️ Eliminar Zonas", "📋 Copiar Zonas"])

    with tab1:
        token = st.text_input("🔑 Token de SimpliRoute", type="password", key="token_zonas")
        uploaded = st.file_uploader("📂 Sube tu archivo KML o RTF", type=["kml","rtf","txt"])
        if uploaded:
            text = decode_file(uploaded.read())
            if text.strip().startswith("{\\rtf") or "\\rtf" in text[:100]:
                text = strip_rtf_codes(text)
            polygons = parse_polygons(text)
            if not polygons:
                st.warning("No se encontraron polígonos."); return
            st.success(f"✅ {len(polygons)} polígono(s)")
            for p in polygons:
                with st.expander(f"📍 {p['name']} — {len(p['coords'])} puntos"):
                    st.code(coords_to_str(p["coords"][:3]) + ",...]")
            st.download_button("⬇️ Descargar Excel", data=generate_excel_zones(polygons), file_name="ZONES.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            st.divider()
            if not token:
                st.warning("⚠️ Ingresa tu token para cargar zonas.")
            elif st.button("🚀 Cargar zonas", type="primary"):
                prog = st.progress(0); status = st.empty()
                for i, p in enumerate(polygons):
                    status.info(f"Cargando: **{p['name']}** ({i+1}/{len(polygons)})")
                    try:
                        r = requests.post("http://api.simpliroute.com/v1/zones/",
                                          headers={"authorization":f"Token {token}","content-type":"application/json"},
                                          json={"name":p["name"],"coordinates":coords_to_str(p["coords"]),"vehicles":[]}, timeout=15)
                        code = r.status_code
                    except: code = None
                    if code in [200,201]: st.success(f"✅ {p['name']}")
                    elif code == 400: st.warning(f"⚠️ {p['name']} — Ya existe")
                    elif code == 401: st.error(f"❌ {p['name']} — Token inválido")
                    else: st.error(f"❌ {p['name']} — Error {code}")
                    prog.progress((i+1)/len(polygons))
                status.empty(); prog.empty()

    with tab3:
        st.markdown("Consulta las zonas de una cuenta origen y cópialas a una cuenta destino.")
        col1, col2 = st.columns(2)
        with col1: token_origen = st.text_input("🔑 Token cuenta origen", type="password", key="token_zonas_origen")
        with col2: token_destino = st.text_input("🔑 Token cuenta destino", type="password", key="token_zonas_destino")

        if st.button("🔍 Consultar zonas", type="primary", disabled=not token_origen, key="btn_consultar_zonas"):
            try:
                r = requests.get("http://api.simpliroute.com/v1/zones/",
                                 headers={"authorization": f"Token {token_origen}", "content-type": "application/json"}, timeout=30)
                if r.status_code == 401: st.error("❌ Token inválido"); return
                if r.status_code != 200: st.error(f"❌ Error {r.status_code}"); return
                st.session_state["zonas_origen"] = r.json()
            except Exception as e:
                st.error(f"❌ Error: {e}"); return

        if "zonas_origen" in st.session_state:
            zonas = st.session_state["zonas_origen"]
            st.success(f"✅ {len(zonas)} zona(s) encontradas")
            st.divider()
            st.markdown("**Selecciona las zonas a copiar:**")

            seleccionadas = []
            cols = st.columns(2)
            for i, zona in enumerate(zonas):
                with cols[i % 2]:
                    if st.checkbox(f"📍 {zona['name']}", key=f"zona_check_{zona['id']}"):
                        seleccionadas.append(zona)

            st.divider()
            if seleccionadas:
                st.info(f"**{len(seleccionadas)} zona(s) seleccionadas** para copiar")

            if st.button("🚀 Copiar zonas", type="primary",
                         disabled=not (token_destino and seleccionadas), key="btn_copiar_zonas"):
                prog = st.progress(0); status = st.empty()
                ok_count = 0; err_count = 0
                for i, zona in enumerate(seleccionadas):
                    status.info(f"Copiando: **{zona['name']}** ({i+1}/{len(seleccionadas)})")
                    coords = zona.get("coordinates", [])
                    coords_str = "[" + ",".join(
                        f"{{'lat':'{p['lat']}','lng':'{p['lng']}'}}" for p in coords
                    ) + "]"
                    try:
                        r = requests.post("http://api.simpliroute.com/v1/zones/",
                                          headers={"authorization": f"Token {token_destino}", "content-type": "application/json"},
                                          json={"name": zona["name"], "coordinates": coords_str, "vehicles": []}, timeout=15)
                        code = r.status_code
                    except Exception as e:
                        st.error(f"❌ **{zona['name']}** — Error: {e}"); err_count += 1
                        prog.progress((i+1)/len(seleccionadas)); continue
                    if code in [200, 201]: st.success(f"✅ {zona['name']}"); ok_count += 1
                    elif code == 400: st.warning(f"⚠️ {zona['name']} — Ya existe")
                    elif code == 401: st.error("❌ Token destino inválido."); status.empty(); prog.empty(); return
                    else: st.error(f"❌ {zona['name']} — Error {code}"); err_count += 1
                    prog.progress((i+1)/len(seleccionadas))
                status.empty(); prog.empty()
                st.divider()
                if err_count == 0: st.success(f"✅ Completado — **{ok_count} zonas copiadas**")
                else: st.warning(f"⚠️ Completado — **{ok_count} copiadas**, {err_count} con error") y no se puede deshacer.")
        token_del = st.text_input("🔑 Token de SimpliRoute", type="password", key="token_zonas_del")
        st.caption("Pega los IDs de zona uno por línea:\n```\n12345\n12346\n```")
        zone_ids_raw = st.text_area("IDs de zona", placeholder="12345\n12346", height=180,
                                    label_visibility="collapsed", key="zone_ids_del")
        confirm_del = st.checkbox("✅ Confirmo que quiero eliminar estas zonas de forma permanente", key="confirm_zonas_del")
        if st.button("🗑️ Eliminar Zonas", type="primary", disabled=not (token_del and zone_ids_raw and confirm_del)):
            zone_ids = []
            for l in zone_ids_raw.strip().splitlines():
                l = l.strip()
                if not l: continue
                try: zone_ids.append(int(l))
                except: st.warning(f"⚠️ ID inválido ignorado: {l}")
            if not zone_ids: st.error("❌ No se encontraron IDs válidos."); return
            prog = st.progress(0); status = st.empty()
            ok_count = 0; err_count = 0
            for i, zid in enumerate(zone_ids):
                status.info(f"Eliminando zona **{zid}** ({i+1}/{len(zone_ids)})...")
                try:
                    r = requests.delete(f"http://api.simpliroute.com/v1/zones/{zid}",
                                        headers={"authorization": f"Token {token_del}", "accept": "application/json"}, timeout=30)
                    code = r.status_code
                except Exception as e:
                    st.error(f"❌ **{zid}** — Error: {e}"); err_count += 1; prog.progress((i+1)/len(zone_ids)); continue
                if code in [200, 204]: st.success(f"✅ Zona **{zid}** eliminada"); ok_count += 1
                elif code == 401: st.error("❌ Token inválido."); status.empty(); prog.empty(); return
                elif code == 404: st.warning(f"⚠️ **{zid}** — No encontrada")
                else: st.error(f"❌ **{zid}** — Error {code}"); err_count += 1
                prog.progress((i+1)/len(zone_ids))
            status.empty(); prog.empty()
            st.divider()
            if err_count == 0: st.success(f"✅ Completado — **{ok_count} zonas eliminadas**")
            else: st.warning(f"⚠️ Completado con errores — **{ok_count} eliminadas**, {err_count} con error")


# ── FEATURE: CAMBIAR ROL ──────────────────────────────────────────────────────
ROLES = {"Administrador":"is_admin","Conductor":"is_driver","Co-conductor":"is_codriver",
         "Router Jr":"is_router_jr","Monitor":"is_monitor","Coordinador":"is_coordinator",
         "Router":"is_router","Staff":"is_staff","Seller Viewer":"is_seller_viewer","Seller":"is_seller"}
ALL_ROLE_KEYS = list(ROLES.values()) + ["is_owner"]

def get_current_role(user):
    for label, key in ROLES.items():
        if user.get(key): return label, key
    if user.get("is_owner"): return "Owner", "is_owner"
    return "Sin rol", None

def page_cambiar_rol():
    st.title("👤 Cambiar Rol de Usuario")
    col1, col2 = st.columns(2)
    with col1: token = st.text_input("🔑 Token", type="password", key="token_rol")
    with col2: user_id = st.text_input("🆔 ID del Usuario", placeholder="Ej: 524614")
    if st.button("🔍 Consultar", type="primary", disabled=not (token and user_id)):
        try:
            r = requests.get(f"http://api.simpliroute.com/v1/accounts/users/{user_id.strip()}/",
                             headers={"Authorization":f"Token {token}"}, timeout=15)
            code, resp = r.status_code, r.json()
        except Exception as e:
            code, resp = None, str(e)
        if code == 200:
            st.session_state["user_data"] = resp; st.session_state["user_token"] = token
        else:
            st.error(f"❌ Error {code}: {resp}"); st.session_state.pop("user_data", None)
    if "user_data" in st.session_state:
        user = st.session_state["user_data"]
        cl, ck = get_current_role(user)
        st.divider()
        c1,c2,c3 = st.columns(3)
        c1.markdown(f"**Nombre**<br>{user.get('name','—')}", unsafe_allow_html=True)
        c2.markdown(f"**Username**<br>{user.get('username','—')}", unsafe_allow_html=True)
        c3.markdown(f"**Rol actual**<br>{cl}", unsafe_allow_html=True)
        st.caption(f"📧 {user.get('email','—')} · {user.get('status','—')}")
        st.divider()
        if ck == "is_owner":
            st.warning("⚠️ Este usuario es Owner. No se puede modificar.")
        else:
            nuevo_label = st.selectbox("Nuevo rol:", [l for l, k in ROLES.items() if k != ck])
            nuevo_key = ROLES[nuevo_label]
            if st.button(f"💾 Cambiar a {nuevo_label}", type="primary"):
                payload = {k: (k == nuevo_key) for k in ALL_ROLE_KEYS}
                payload["username"] = user.get("username",""); payload["name"] = user.get("name","")
                try:
                    r = requests.put(f"http://api.simpliroute.com/v1/accounts/users/{user['id']}/",
                                     headers={"Authorization":f"Token {st.session_state['user_token']}","Content-Type":"application/json"},
                                     json=payload, timeout=15)
                    code = r.status_code
                except: code = None
                if code == 200:
                    st.success(f"✅ Rol actualizado a **{nuevo_label}**")
                    try:
                        r2 = requests.get(f"http://api.simpliroute.com/v1/accounts/users/{user['id']}/",
                                          headers={"Authorization":f"Token {st.session_state['user_token']}"}, timeout=15)
                        if r2.status_code == 200: st.session_state["user_data"] = r2.json()
                    except: pass
                    st.rerun()
                else: st.error(f"❌ Error {code}")


# ── FEATURE: CREAR WEBHOOK ────────────────────────────────────────────────────
WEBHOOK_EVENTS = {"Plan creado":"plan_created","Plan editado":"plan_edited","Ruta creada":"route_created",
                  "Ruta editada":"route_edited","Ruta iniciada":"route_started","Ruta finalizada":"route_finished",
                  "En camino":"on_its_way","Checkout":"visit_checkout","Checkout detallado":"visit_checkout_detailed"}

def page_crear_webhook():
    st.title("🔔 Crear Webhook")
    token = st.text_input("🔑 Token de SimpliRoute", type="password", key="token_webhook")
    url = st.text_input("🌐 URL de destino", placeholder="https://tu-servidor.com/webhook")
    evento_key = WEBHOOK_EVENTS[st.selectbox("📋 Tipo de evento", list(WEBHOOK_EVENTS.keys()))]
    custom = st.checkbox("Usar header personalizado")
    if custom:
        raw = st.text_area("Header JSON", value='{\n    "Content-Type": "application/json",\n    "X-Custom": "valor"\n}',
                           height=150, label_visibility="collapsed")
        try: headers_payload = json.loads(raw)
        except: st.error("❌ JSON inválido"); headers_payload = None
    else:
        headers_payload = {"Content-Type": "application/json"}
        st.code('{"Content-Type": "application/json"}', language="json")
    st.divider()
    if st.button("🚀 Crear Webhook", type="primary", disabled=not (token and url)):
        if not url.startswith("http"): st.error("❌ URL inválida"); return
        if custom and not headers_payload: st.error("❌ Corrige el JSON"); return
        try:
            r = requests.post("http://api.simpliroute.com/v1/addons/webhooks/",
                              headers={"Authorization":f"Token {token}","Content-Type":"application/json"},
                              json={"webhook":evento_key,"url":url,"headers":headers_payload}, timeout=30)
            code, resp = r.status_code, r.json()
        except Exception as e:
            code, resp = None, str(e)
        if code in [200,201]: st.success("✅ Webhook creado"); st.json(resp)
        elif code == 401: st.error("❌ Token inválido")
        else: st.error(f"❌ Error {code}: {resp}")


# ── FEATURE: DESBLOQUEO DE CONTRASEÑA ────────────────────────────────────────
SUPPORT_AGENTS = {
    "Brandon Vargas": "brandon.vargas@simpliroute.com",
    "Carlos Junior": "carlos.celestino@simpliroute.com",
    "David Martinez": "david.martinez@simpliroute.com",
    "Itzel Meza": "itzel.meza@simpliroute.com",
    "Jorge Cruz": "jorge.cruz@simpliroute.com",
    "Julio Mares": "julio.mares@simpliroute.com",
    "Roger Camacho": "roger.camacho@simpliroute.com",
    "Silmary Guedez": "silmary.guedez@simpliroute.com",
}

def page_desbloqueo():
    st.title("🔓 Desbloqueo de Contraseña")
    token = st.text_input("🔑 Token de SimpliRoute", type="password", key="token_unlock")
    if st.button("🔍 Consultar usuarios bloqueados", type="primary", disabled=not token):
        with st.spinner("Consultando..."):
            code, resp = get_users_list(token)
        if code == 200:
            blocked = [u for u in resp if u.get("blocked") is True]
            if not blocked:
                st.info("✅ No hay usuarios bloqueados.")
                st.session_state.pop("blocked_users", None)
            else:
                st.session_state["blocked_users"] = blocked
                st.session_state["unlock_token"] = token
        else:
            st.error(f"❌ Error {code}")

    if "pending_restore" in st.session_state:
        pr = st.session_state["pending_restore"]
        st.divider()
        st.subheader("📧 Restauración pendiente")
        st.info(f"Usuario: **{pr['user'].get('username')}** · Email original: **{pr['original_email'] or 'Sin email'}**")
        st.markdown("Una vez que el usuario haya reseteado su contraseña, presiona el botón para restaurar su email original.")
        if st.button("↩️ Confirmar y restaurar email original", type="primary", key="btn_restore"):
            code, _ = put_user_full(pr["token"], pr["user"], pr["original_email"])
            if code in [200,201]:
                st.success("✅ Email restaurado."); st.session_state.pop("pending_restore", None)
            else:
                st.error(f"❌ Error {code}")

    if "blocked_users" in st.session_state:
        blocked = st.session_state["blocked_users"]
        st.divider()
        st.success(f"⚠️ {len(blocked)} usuario(s) bloqueado(s)")
        def ulabel(u):
            name = u.get("name") or u.get("username") or str(u.get("id"))
            un = u.get("username","")
            return f"{name} — {un}" if un != name else name
        selected_user = {ulabel(u): u for u in blocked}[st.selectbox("Usuario bloqueado:", [ulabel(u) for u in blocked])]
        original_email = selected_user.get("email","")
        st.caption(f"📧 Email actual: **{original_email or 'Sin email'}**")
        st.divider()
        agent_name = st.selectbox("Agente que recibirá el link:", list(SUPPORT_AGENTS.keys()))
        recovery_email = SUPPORT_AGENTS[agent_name]
        st.caption(f"📨 Se enviará a: **{recovery_email}**")
        st.divider()
        if st.button("🔓 Enviar link de desbloqueo", type="primary"):
            tok = st.session_state["unlock_token"]
            prog = st.progress(0); status = st.empty()
            status.info("1/2 — Asignando email temporal...")
            code1, _ = put_user_full(tok, selected_user, recovery_email)
            prog.progress(50)
            if code1 not in [200,201]:
                st.error(f"❌ Error al actualizar email: {code1}"); status.empty(); prog.empty()
            else:
                status.info("2/2 — Enviando link...")
                try:
                    r = requests.post("https://api.simpliroute.com/v2/auth/unlock/",
                                      headers={"Content-Type":"application/json;charset=UTF-8","authorization":"null"},
                                      json={"username": selected_user.get("username")}, timeout=30)
                    code2 = r.status_code
                except: code2 = None
                prog.progress(100); status.empty(); prog.empty()
                if code2 in [200,201]:
                    st.success(f"✅ Link enviado a **{recovery_email}** ({agent_name})")
                    st.warning("⚠️ Cuando el usuario resetee su contraseña, usa el botón de restauración de arriba.")
                    st.session_state["pending_restore"] = {"user": selected_user, "original_email": original_email, "token": tok}
                    st.session_state.pop("blocked_users", None)
                    st.rerun()
                else:
                    put_user_full(tok, selected_user, original_email)
                    st.error(f"❌ No se pudo enviar el link. Email restaurado. ({code2})")


# ── FEATURE: EDICIÓN DE VISITAS ───────────────────────────────────────────────
def page_edicion_visitas():
    st.title("✏️ Edición de Visitas")
    st.markdown("Edita la fecha planificada o la ruta asignada de un lote de visitas.")
    token = st.text_input("🔑 Token de SimpliRoute", type="password", key="token_edit_visits")
    st.divider()
    st.subheader("📋 IDs de visita")
    st.caption("Pega los IDs uno por línea:\n```\n833673298\n837739792\n```")
    visit_ids_raw = st.text_area("IDs de visita", placeholder="833673298\n837739792", height=180, label_visibility="collapsed")
    st.divider()
    st.subheader("📅 Fecha planificada")
    date_action = st.radio("Acción sobre la fecha:", ["No cambiar","Asignar nueva fecha","Eliminar fecha"], horizontal=True, key="date_action")
    new_date = None
    if date_action == "Asignar nueva fecha":
        new_date = st.date_input("Selecciona la fecha:", value=date.today(), key="edit_date")
        st.caption(f"Se asignará: **{new_date.strftime('%Y-%m-%d')}**")
    elif date_action == "Eliminar fecha":
        st.warning("⚠️ Se eliminará la fecha planificada de todas las visitas.")
    st.divider()
    st.subheader("🛣️ Ruta")
    route_action = st.radio("Acción sobre la ruta:", ["No cambiar","Asignar ruta","Eliminar ruta"], horizontal=True, key="route_action")
    new_route = None
    if route_action == "Asignar ruta":
        new_route = st.text_input("ID de la ruta:", placeholder="5859a5d1-03c5-4152-bcea-500bab2ad47d")
    elif route_action == "Eliminar ruta":
        st.warning("⚠️ Se eliminará la ruta asignada de todas las visitas.")
    st.divider()
    no_changes = date_action == "No cambiar" and route_action == "No cambiar"
    if st.button("🚀 Editar visitas", type="primary", disabled=not (token and visit_ids_raw) or no_changes):
        visit_ids = []
        for l in visit_ids_raw.strip().splitlines():
            l = l.strip()
            if not l: continue
            try: visit_ids.append(int(l))
            except: st.warning(f"⚠️ ID inválido ignorado: {l}")
        if not visit_ids: st.error("❌ No se encontraron IDs válidos."); return

        def build_payload(vid):
            p = {"id": vid}
            if date_action == "Asignar nueva fecha": p["planned_date"] = new_date.strftime("%Y-%m-%d")
            elif date_action == "Eliminar fecha": p["planned_date"] = None
            if route_action == "Asignar ruta": p["route"] = new_route.strip()
            elif route_action == "Eliminar ruta": p["route"] = None
            return p

        prog = st.progress(0); status = st.empty()
        ok_count = 0; err_count = 0
        batches = [visit_ids[i:i+500] for i in range(0, len(visit_ids), 500)]
        for i, batch in enumerate(batches):
            status.info(f"Editando lote {i+1}/{len(batches)} — {len(batch)} visitas... ({ok_count} editadas)")
            try:
                r = requests.patch("http://api.simpliroute.com/v1/routes/visits/",
                                   headers={"Authorization":f"Token {token}","Content-Type":"application/json"},
                                   json=[build_payload(vid) for vid in batch], timeout=300)
                code = r.status_code
            except Exception as e:
                if "timed out" in str(e).lower(): st.warning(f"⏱️ **Lote {i+1}** — Tiempo de espera agotado.")
                else: st.error(f"❌ Error en lote {i+1}: {e}")
                err_count += 1; prog.progress((i+1)/len(batches)); continue
            if code in [200,201]:
                ok_count += len(batch); st.success(f"✅ Lote {i+1}/{len(batches)} — {len(batch)} visitas editadas")
            elif code == 401:
                st.error("❌ Token inválido."); status.empty(); prog.empty(); return
            else:
                st.error(f"❌ Lote {i+1} — Error {code}"); err_count += 1
            prog.progress((i+1)/len(batches))
        status.empty(); prog.empty()
        st.divider()
        if err_count == 0: st.success(f"✅ Completado — **{ok_count} visitas editadas**")
        else: st.warning(f"⚠️ Completado con errores — **{ok_count} editadas**, {err_count} lote(s) fallido(s)")


# ── FEATURE: ELIMINACIÓN MASIVA DE VISITAS ───────────────────────────────────
def page_eliminacion_visitas():
    st.title("🗑️ Eliminación Masiva de Visitas")
    st.error("⚠️ **ADVERTENCIA:** Esta operación elimina visitas directamente desde la base de datos. La eliminación es permanente y no se puede deshacer.")
    token = st.text_input("🔑 Token de SimpliRoute", type="password", key="token_delete")
    def make_template():
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Visitas"
        ws.append(["id"]); ws.append([838112279]); ws.append([838112568])
        ws.column_dimensions["A"].width = 20
        buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf
    st.download_button("📥 Descargar plantilla", data=make_template(), file_name="plantilla_eliminacion_visitas.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    visits_file = st.file_uploader("📂 Sube tu Excel con IDs de visita", type=["xlsx"], key="upload_delete")
    if visits_file:
        try:
            wb = openpyxl.load_workbook(visits_file); ws = wb.active
            headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
            if "id" not in headers: st.error("❌ El Excel debe tener columna 'id'."); return
            idx = headers.index("id")
            visit_ids = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                val = row[idx] if idx < len(row) else None
                if val is not None:
                    try: visit_ids.append(int(val))
                    except: pass
            if not visit_ids: st.error("❌ No se encontraron IDs."); return
            st.info(f"📋 **{len(visit_ids)} visitas** cargadas. Se procesarán en lotes de 2,000.")
        except Exception as e:
            st.error(f"❌ Error: {e}"); return
        confirm = st.checkbox("✅ Confirmo que quiero eliminar estas visitas de forma permanente")
        if confirm and st.button("🗑️ Eliminar visitas", type="primary", disabled=not token):
            batches = [visit_ids[i:i+2000] for i in range(0, len(visit_ids), 2000)]
            prog = st.progress(0); status = st.empty()
            total_deleted = 0; errors = 0
            for i, batch in enumerate(batches):
                status.info(f"Procesando lote {i+1}/{len(batches)} — {len(batch)} visitas... ({total_deleted} eliminadas)")
                try:
                    r = requests.post("http://api.simpliroute.com/v1/bulk/delete/visits/",
                                      headers={"Authorization":f"Token {token}","Content-Type":"application/json"},
                                      json={"visits": batch}, timeout=300)
                    code = r.status_code
                except Exception as e:
                    if "timed out" in str(e).lower(): st.warning(f"⏱️ **Lote {i+1}** — Tiempo de espera agotado.")
                    else: st.error(f"❌ Error en lote {i+1}: {e}")
                    errors += 1; prog.progress((i+1)/len(batches)); continue
                if code in [200,201,204]:
                    total_deleted += len(batch); st.success(f"✅ Lote {i+1}/{len(batches)} — {len(batch)} visitas eliminadas")
                elif code == 401:
                    st.error("❌ Token inválido."); status.empty(); prog.empty(); return
                else:
                    st.error(f"❌ Lote {i+1} — Error {code}"); errors += 1
                prog.progress((i+1)/len(batches))
            status.empty(); prog.empty()
            st.divider()
            if errors == 0: st.success(f"✅ Completado — **{total_deleted} visitas eliminadas**")
            else: st.warning(f"⚠️ Completado con errores — **{total_deleted} eliminadas**, {errors} lote(s) fallido(s)")


# ── FEATURE: INICIAR / CERRAR RUTAS ──────────────────────────────────────────
def page_iniciar_cerrar_rutas():
    st.title("🚦 Iniciar / Cerrar Rutas")
    token = st.text_input("🔑 Token de SimpliRoute", type="password", key="token_routes")
    evento = st.selectbox("📋 Tipo de evento", ["Iniciar ruta","Finalizar ruta"])
    event_type = "ROUTE_STARTED" if evento == "Iniciar ruta" else "ROUTE_FINISHED"
    selected_date = st.date_input("📅 Fecha", value=date.today(), key="route_date")
    hora = "12:00:00.000Z" if event_type == "ROUTE_STARTED" else "22:00:00.000Z"
    date_time = f"{selected_date.strftime('%Y-%m-%d')}T{hora}"
    st.caption(f"📅 Fecha y hora que se usará: **{date_time}**")
    st.divider()
    st.caption("Pega los IDs de ruta uno por línea:\n```\n637f11a2-a1a6-4609-8c23-83e8c76dccbf\n```")
    route_ids_raw = st.text_area("IDs de ruta", placeholder="637f11a2-a1a6-4609-8c23-83e8c76dccbf", height=200, label_visibility="collapsed")
    if st.button(f"🚀 {evento}", type="primary", disabled=not (token and route_ids_raw)):
        route_ids = [l.strip() for l in route_ids_raw.strip().splitlines() if l.strip()]
        if not route_ids: st.error("❌ No se encontraron IDs."); return
        prog = st.progress(0); status = st.empty()
        for i, route_id in enumerate(route_ids):
            status.info(f"Procesando: **{route_id}** ({i+1}/{len(route_ids)})")
            try:
                r = requests.get(f"http://api.simpliroute.com/v1/routes/routes/{route_id}/",
                                 headers={"Authorization":f"Token {token}","Content-Type":"application/json"}, timeout=30)
                code_get, route_data = r.status_code, r.json()
            except Exception as e:
                st.error(f"❌ **{route_id}** — Error: {e}"); prog.progress((i+1)/len(route_ids)); continue
            if code_get != 200:
                st.error(f"❌ **{route_id}** — Error {code_get}"); prog.progress((i+1)/len(route_ids)); continue
            lat = route_data.get("location_start_latitude") if event_type == "ROUTE_STARTED" else route_data.get("location_end_latitude")
            lng = route_data.get("location_start_longitude") if event_type == "ROUTE_STARTED" else route_data.get("location_end_longitude")
            if not lat or not lng:
                st.warning(f"⚠️ **{route_id}** — Sin coordenadas"); prog.progress((i+1)/len(route_ids)); continue
            try:
                r2 = requests.post("http://api.simpliroute.com/v1/events/register/",
                                   headers={"Authorization":f"Token {token}","Content-Type":"application/json"},
                                   json={"date_time":date_time,"latitude":float(lat),"longitude":float(lng),
                                         "route_id":route_id,"type":event_type}, timeout=30)
                code_post = r2.status_code
            except Exception as e:
                st.error(f"❌ **{route_id}** — Error: {e}"); prog.progress((i+1)/len(route_ids)); continue
            if code_post in [200,201]: st.success(f"✅ **{route_id}** — {evento} registrado")
            elif code_post == 401: st.error(f"❌ **{route_id}** — Token inválido")
            else: st.error(f"❌ **{route_id}** — Error {code_post}")
            prog.progress((i+1)/len(route_ids))
        status.empty(); prog.empty()


# ── FEATURE: ANÁLISIS DE RECORRIDO GPS ───────────────────────────────────────
def page_analisis_gps():
    st.title("📍 Análisis de Recorrido GPS")
    st.markdown("Carga un JSON de puntos GPS para analizar el recorrido, detectar anomalías y comparar contra un punto fijo.")
    json_file = st.file_uploader("📂 Sube tu archivo JSON", type=["json"], key="upload_gps_analysis")
    if not json_file: return
    try:
        data = json.loads(json_file.read())
        if not isinstance(data, list) or len(data) < 2:
            st.error("❌ El JSON debe ser una lista con al menos 2 puntos."); return
    except Exception as e:
        st.error(f"❌ Error leyendo el JSON: {e}"); return
    st.success(f"✅ {len(data)} puntos cargados")
    st.divider()
    st.subheader("🛣️ Análisis de recorrido")
    st.caption("Puntos anómalos: velocidad implícita entre puntos consecutivos **> 500 km/h**.")
    VELOCIDAD_MAX = 500
    total_bruto = 0; total_limpio = 0; anomalos = []; detalles = []
    for i in range(1, len(data)):
        p1, p2 = data[i-1], data[i]
        try:
            lat1, lon1 = float(p1["latitude"]), float(p1["longitude"])
            lat2, lon2 = float(p2["latitude"]), float(p2["longitude"])
            dist_m = haversine_m(lat1, lon1, lat2, lon2)
            velocidad_kmh = None
            try:
                t1 = datetime.fromisoformat(p1["timestamp"].replace("Z","+00:00"))
                t2 = datetime.fromisoformat(p2["timestamp"].replace("Z","+00:00"))
                seg = abs((t2-t1).total_seconds())
                if seg > 0: velocidad_kmh = (dist_m/1000)/(seg/3600)
            except: pass
            es_anomalo = velocidad_kmh is not None and velocidad_kmh > VELOCIDAD_MAX
            total_bruto += dist_m
            if not es_anomalo: total_limpio += dist_m
            detalles.append({"index":i,"timestamp":p2.get("timestamp","—"),"lat":lat2,"lon":lon2,
                             "dist_m":round(dist_m,2),"velocidad_kmh":round(velocidad_kmh,1) if velocidad_kmh else None,"anomalo":es_anomalo})
            if es_anomalo:
                anomalos.append({"index":i,"timestamp":p2.get("timestamp","—"),"lat":lat2,"lon":lon2,
                                 "dist_km":round(dist_m/1000,2),"velocidad_kmh":round(velocidad_kmh,1)})
        except: continue
    col1,col2,col3 = st.columns(3)
    col1.metric("📏 Total bruto", f"{total_bruto/1000:.2f} km")
    col2.metric("✅ Total limpio", f"{total_limpio/1000:.2f} km")
    col3.metric("⚠️ Puntos anómalos", len(anomalos))
    if anomalos:
        top10 = sorted(anomalos, key=lambda x: x["dist_km"], reverse=True)[:10]
        st.markdown(f"**{len(anomalos)} puntos anómalos — Top 10:**")
        for a in top10:
            st.error(f"⚠️ Índice {a['index']} | `{a['timestamp']}` | `{a['lat']}, {a['lon']}` | {a['dist_km']} km | {a['velocidad_kmh']} km/h")
        def make_excel_anomalos(items):
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Anomalos"
            ws.append(["index","timestamp","latitude","longitude","distancia_km","velocidad_kmh"])
            for a in sorted(items, key=lambda x: x["dist_km"], reverse=True):
                ws.append([a["index"],a["timestamp"],a["lat"],a["lon"],a["dist_km"],a["velocidad_kmh"]])
            ws.column_dimensions["B"].width = 25
            buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf
        st.download_button("⬇️ Descargar Excel anomalos", data=make_excel_anomalos(anomalos),
                           file_name="puntos_anomalos.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    else:
        st.success("✅ No se detectaron puntos anómalos.")
    def make_excel_analisis(detalles):
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Recorrido"
        ws.append(["index","timestamp","latitude","longitude","distancia_m","velocidad_kmh","anomalo"])
        for d in detalles:
            ws.append([d["index"],d["timestamp"],d["lat"],d["lon"],d["dist_m"],d["velocidad_kmh"],"Sí" if d["anomalo"] else "No"])
        ws.column_dimensions["B"].width = 25
        buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf
    st.download_button("⬇️ Descargar Excel recorrido", data=make_excel_analisis(detalles),
                       file_name="analisis_recorrido.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    st.divider()
    st.subheader("📌 Comparación con punto fijo")
    col_a,col_b,col_c = st.columns(3)
    with col_a: ref_lat = st.text_input("🌐 Latitud", placeholder="-22.797010")
    with col_b: ref_lon = st.text_input("🌐 Longitud", placeholder="-43.323240")
    with col_c: radio_m = st.number_input("📏 Radio (metros)", min_value=1, value=500, step=50)
    if st.button("🔍 Comparar puntos", type="primary", disabled=not (ref_lat and ref_lon)):
        try: rlat, rlon = float(ref_lat), float(ref_lon)
        except: st.error("❌ Coordenadas inválidas."); return
        dentro = []; fuera = 0
        for i, p in enumerate(data):
            try:
                plat, plon = float(p["latitude"]), float(p["longitude"])
                dist = haversine_m(rlat, rlon, plat, plon)
                if dist <= radio_m: dentro.append({"index":i,"timestamp":p.get("timestamp","—"),"lat":plat,"lon":plon,"dist_m":round(dist,1)})
                else: fuera += 1
            except: continue
        col_x,col_y = st.columns(2)
        col_x.metric(f"✅ Dentro ({radio_m}m)", len(dentro))
        col_y.metric("❌ Fuera", fuera)
        if dentro:
            with st.expander(f"Ver {len(dentro)} puntos"):
                for p in dentro[:50]:
                    st.markdown(f"- Índice `{p['index']}` | `{p['timestamp']}` | `{p['lat']}, {p['lon']}` | **{p['dist_m']} m**")
                if len(dentro) > 50: st.caption(f"... y {len(dentro)-50} más.")
            def make_excel_radio(puntos):
                wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Puntos cercanos"
                ws.append(["index","timestamp","latitude","longitude","distancia_al_punto_m"])
                for p in puntos: ws.append([p["index"],p["timestamp"],p["lat"],p["lon"],p["dist_m"]])
                ws.column_dimensions["B"].width = 25
                buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf
            st.download_button("⬇️ Descargar Excel", data=make_excel_radio(dentro),
                               file_name=f"puntos_dentro_{radio_m}m.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        else:
            st.info(f"ℹ️ Ningún punto dentro de {radio_m} metros.")


# ── FEATURE: REENVIAR WEBHOOKS ────────────────────────────────────────────────
def page_reenviar_webhooks():
    st.title("🔁 Reenviar Webhooks")
    st.info("🔑 Usa el token de tu cuenta de SimpliRoute.")
    token = st.text_input("🔑 Token", type="password", key="token_resend")
    account_id = st.text_input("🏢 ID de la cuenta", placeholder="Ej: 30610")
    st.caption("Pega los IDs de visita uno por línea:\n```\n799841373\n808472905\n```")
    visit_ids_raw = st.text_area("IDs", placeholder="799841373\n808472905", height=200, label_visibility="collapsed")
    today = date.today().strftime("%Y-%m-%d")
    st.caption(f"📅 Fecha: **{today}**")
    st.divider()
    if st.button("🚀 Reenviar", type="primary", disabled=not (token and account_id and visit_ids_raw)):
        ids, errs = [], []
        for l in visit_ids_raw.strip().splitlines():
            l = l.strip()
            try: ids.append(int(l))
            except: errs.append(l)
        if errs: st.error(f"❌ IDs inválidos: {', '.join(errs)}"); return
        with st.spinner(f"Reenviando {len(ids)} visitas..."):
            try:
                r = requests.post("https://api.simpliroute.com/v1/mobile/send-webhooks",
                                  headers={"Authorization":f"Token {token}","Content-Type":"application/json"},
                                  json={"account_ids":[int(account_id)],"planned_date":today,"visit_ids":ids}, timeout=60)
                code = r.status_code
            except: code = None
        if code in [200,201]: st.success(f"✅ Webhooks reenviados para {len(ids)} visitas")
        elif code == 401: st.error("❌ Token inválido")
        elif code == 404: st.error("❌ Cuenta no encontrada")
        else: st.error(f"❌ Error {code}")


# ── FEATURE: TIPOS DE VISITA Y SKILLS ────────────────────────────────────────
def page_visit_types_skills():
    st.title("🏷️ Tipos de Visita y Skills")
    token = st.text_input("🔑 Token de SimpliRoute", type="password", key="token_vts")
    st.markdown("---"); st.subheader("🔧 Skills")
    def tpl_skills():
        wb=openpyxl.Workbook(); ws=wb.active; ws.title="Skills"
        ws.append(["skill"]); ws.append(["Manejo de carga pesada"]); ws.append(["Refrigeración"])
        ws.column_dimensions["A"].width=35; buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf
    st.download_button("📥 Plantilla Skills", data=tpl_skills(), file_name="plantilla_skills.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    sf = st.file_uploader("📂 Excel Skills", type=["xlsx"], key="upload_skills")
    if sf:
        try:
            rows = read_excel_column(sf, ["skill"])
            st.success(f"✅ {len(rows)} skill(s)")
            with st.expander("Ver"): [st.markdown(f"- {r['skill']}") for r in rows]
        except Exception as e:
            rows = []; st.error(str(e))
        if rows and token and st.button("🚀 Crear Skills", type="primary", key="btn_skills"):
            results=[]; prog=st.progress(0); status=st.empty()
            for i, row in enumerate(rows):
                status.info(f"Creando: **{row['skill']}** ({i+1}/{len(rows)})")
                try:
                    r=requests.post("http://api.simpliroute.com/v1/routes/skills/",
                                    headers={"Authorization":f"Token {token}","Content-Type":"application/json"},
                                    json={"skill":row["skill"]}, timeout=15)
                    code,resp=r.status_code,r.json()
                except Exception as e: code,resp=None,str(e)
                results.append({"skill":row["skill"],"code":code,"resp":resp}); prog.progress((i+1)/len(rows))
            status.empty(); prog.empty(); show_results(results, "skill")
    st.markdown("---"); st.subheader("📋 Tipos de Visita")
    def tpl_vt():
        wb=openpyxl.Workbook(); ws=wb.active; ws.title="Tipos de Visita"
        ws.append(["label","key"]); ws.append(["Entrega express","entrega_express"]); ws.append(["Retiro","retiro"])
        ws.column_dimensions["A"].width=30; ws.column_dimensions["B"].width=30
        buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf
    st.download_button("📥 Plantilla Tipos de Visita", data=tpl_vt(), file_name="plantilla_tipos_visita.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    vf = st.file_uploader("📂 Excel Tipos de Visita", type=["xlsx"], key="upload_vt")
    if vf:
        try:
            rows_vt = read_excel_column(vf, ["label","key"])
            st.success(f"✅ {len(rows_vt)} tipo(s)")
            with st.expander("Ver"): [st.markdown(f"- **{r['label']}** → `{r['key']}`") for r in rows_vt]
        except Exception as e:
            rows_vt=[]; st.error(str(e))
        if rows_vt and token and st.button("🚀 Crear Tipos de Visita", type="primary", key="btn_vt"):
            results=[]; prog=st.progress(0); status=st.empty()
            for i, row in enumerate(rows_vt):
                status.info(f"Creando: **{row['label']}** ({i+1}/{len(rows_vt)})")
                try:
                    r=requests.post("http://api.simpliroute.com/v1/accounts/visit-types/",
                                    headers={"Authorization":f"Token {token}","Content-Type":"application/json"},
                                    json=[{"label":row["label"],"key":row["key"]}], timeout=15)
                    code,resp=r.status_code,r.json()
                except Exception as e: code,resp=None,str(e)
                results.append({"label":row["label"],"code":code,"resp":resp}); prog.progress((i+1)/len(rows_vt))
            status.empty(); prog.empty(); show_results(results, "label")


# ── FEATURE: VALIDACIÓN DE GPS ────────────────────────────────────────────────
def page_validacion_gps():
    st.title("📡 Validación de GPS")
    st.markdown("Consulta si un vehículo o conductor tiene registros de ubicación para una fecha determinada.")
    col1, col2 = st.columns(2)
    with col1: token = st.text_input("🔑 Token de SimpliRoute", type="password", key="token_gps")
    with col2: selected_date = st.date_input("📅 Fecha", value=date.today(), key="gps_date")
    date_str = selected_date.strftime("%Y-%m-%d")
    modo = st.radio("🔍 Modo de consulta", ["Por tipo (Vehículo o Conductor)", "Por par (Driver + Vehículo)"], horizontal=True)
    st.divider()
    if modo == "Por tipo (Vehículo o Conductor)":
        entity_type = st.selectbox("Consultar por", ["Vehículo", "Conductor"])
        st.caption("Pega los IDs uno por línea:\n```\n568025\n568026\n```")
        entity_ids_raw = st.text_area("IDs", placeholder="568025\n568026", height=150, label_visibility="collapsed", key="single_ids")
        if st.button("🔍 Consultar GPS", type="primary", disabled=not (token and entity_ids_raw)):
            param_key = "vehicle_id" if entity_type == "Vehículo" else "driver_id"
            entity_ids = [l.strip() for l in entity_ids_raw.strip().splitlines() if l.strip()]
            if not entity_ids: st.error("❌ No se encontraron IDs."); return
            con_data = []; sin_data = []; all_records = []
            prog = st.progress(0); status = st.empty()
            for i, eid in enumerate(entity_ids):
                status.info(f"Consultando {entity_type.lower()} **{eid}** ({i+1}/{len(entity_ids)})")
                url = f"https://api-gateway.simpliroute.com/v1/tracking/locations/{date_str}/?{param_key}={eid}"
                try:
                    r = requests.get(url, headers={"Authorization": f"Token {token}"}, timeout=30)
                    code, data = r.status_code, r.json()
                except Exception as e:
                    st.error(f"❌ **{eid}** — Error: {e}"); prog.progress((i+1)/len(entity_ids)); continue
                if code == 401: st.error("❌ Token inválido."); status.empty(); prog.empty(); return
                elif code != 200: st.error(f"❌ **{eid}** — Error {code}")
                elif not data: sin_data.append(eid)
                else:
                    con_data.append({"id": eid, "count": len(data)})
                    for rec in data: rec["entity_id"] = eid; all_records.append(rec)
                prog.progress((i+1)/len(entity_ids))
            status.empty(); prog.empty()
            st.divider()
            col_a, col_b = st.columns(2)
            with col_a:
                st.success(f"✅ Con data GPS: **{len(con_data)}**")
                for item in con_data: st.markdown(f"- ID `{item['id']}` — {item['count']} registros")
            with col_b:
                st.error(f"❌ Sin data GPS: **{len(sin_data)}**")
                for eid in sin_data: st.markdown(f"- ID `{eid}`")
            if all_records:
                def make_excel_single(records):
                    wb=openpyxl.Workbook(); ws=wb.active; ws.title="GPS"
                    ws.append(["entity_id","timestamp","latitude","longitude","activity_type","type","id","accuracy"])
                    for rec in records:
                        ws.append([rec.get("entity_id",""),rec.get("timestamp",""),rec.get("latitude",""),
                                   rec.get("longitude",""),rec.get("activity_type",""),rec.get("type",""),
                                   rec.get("id",""),rec.get("accuracy","")])
                    ws.column_dimensions["B"].width=25
                    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf
                col_dl1, col_dl2 = st.columns(2)
                with col_dl1:
                    st.download_button("⬇️ Descargar Excel", data=make_excel_single(all_records),
                                       file_name=f"gps_{entity_type.lower()}_{date_str}.xlsx",
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                with col_dl2:
                    st.download_button("⬇️ Descargar JSON", data=json.dumps(all_records, ensure_ascii=False, indent=2),
                                       file_name=f"gps_{entity_type.lower()}_{date_str}.json", mime="application/json")
    else:
        col3, col4 = st.columns(2)
        with col3:
            st.caption("ID de conductor:")
            driver_ids_raw = st.text_area("Driver IDs", placeholder="523842", height=80, label_visibility="collapsed", key="driver_ids")
        with col4:
            st.caption("ID de vehículo:")
            vehicle_ids_raw = st.text_area("Vehicle IDs", placeholder="674797", height=80, label_visibility="collapsed", key="vehicle_ids")
        if st.button("🔍 Consultar GPS", type="primary", disabled=not (token and driver_ids_raw and vehicle_ids_raw), key="btn_gps_pair"):
            driver_ids = [l.strip() for l in driver_ids_raw.strip().splitlines() if l.strip()]
            vehicle_ids = [l.strip() for l in vehicle_ids_raw.strip().splitlines() if l.strip()]
            if len(driver_ids) != len(vehicle_ids):
                st.error(f"❌ Cantidad no coincide: {len(driver_ids)} conductores vs {len(vehicle_ids)} vehículos."); return
            con_data = []; sin_data = []; all_records = []
            prog = st.progress(0); status = st.empty()
            for i, (did, vid) in enumerate(zip(driver_ids, vehicle_ids)):
                status.info(f"Consultando conductor **{did}** / vehículo **{vid}** ({i+1}/{len(driver_ids)})")
                url = f"https://api-gateway.simpliroute.com/v1/tracking/locations/{date_str}/?driver_id={did}&vehicle_id={vid}"
                try:
                    r = requests.get(url, headers={"Authorization": f"Token {token}"}, timeout=30)
                    code, data = r.status_code, r.json()
                except Exception as e:
                    st.error(f"❌ **{did}/{vid}** — Error: {e}"); prog.progress((i+1)/len(driver_ids)); continue
                if code == 401: st.error("❌ Token inválido."); status.empty(); prog.empty(); return
                elif code != 200: st.error(f"❌ **{did}/{vid}** — Error {code}")
                elif not data: sin_data.append(f"Driver {did} / Vehículo {vid}")
                else:
                    con_data.append({"driver": did, "vehicle": vid, "count": len(data)})
                    for rec in data: rec["driver_id"] = did; rec["vehicle_id"] = vid; all_records.append(rec)
                prog.progress((i+1)/len(driver_ids))
            status.empty(); prog.empty()
            col_a, col_b = st.columns(2)
            with col_a:
                st.success(f"✅ Con data GPS: **{len(con_data)}**")
                for item in con_data: st.markdown(f"- Driver `{item['driver']}` / Vehículo `{item['vehicle']}` — {item['count']} registros")
            with col_b:
                st.error(f"❌ Sin data GPS: **{len(sin_data)}**")
                for pair in sin_data: st.markdown(f"- {pair}")
            if all_records:
                def make_excel_pair(records):
                    wb=openpyxl.Workbook(); ws=wb.active; ws.title="GPS"
                    ws.append(["driver_id","vehicle_id","timestamp","latitude","longitude","activity_type","type","id","accuracy"])
                    for rec in records:
                        ws.append([rec.get("driver_id",""),rec.get("vehicle_id",""),rec.get("timestamp",""),
                                   rec.get("latitude",""),rec.get("longitude",""),rec.get("activity_type",""),
                                   rec.get("type",""),rec.get("id",""),rec.get("accuracy","")])
                    ws.column_dimensions["C"].width=25
                    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf
                col_dl3, col_dl4 = st.columns(2)
                with col_dl3:
                    st.download_button("⬇️ Descargar Excel", data=make_excel_pair(all_records),
                                       file_name=f"gps_pares_{date_str}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                with col_dl4:
                    st.download_button("⬇️ Descargar JSON", data=json.dumps(all_records, ensure_ascii=False, indent=2),
                                       file_name=f"gps_pares_{date_str}.json", mime="application/json")


# ── TMS: TIPOS DE DOCUMENTO ───────────────────────────────────────────────────
def page_tms_document_types():
    st.title("📄 Tipos de Documento TMS")
    COUNTRY_OPTIONS = ["PE","CL","MX","CO","AR","EC","BO","UY","PY","VE"]
    c1,c2,c3 = st.columns(3)
    with c1: token = st.text_input("🔑 Token", type="password", key="token_doctype")
    with c2: account_id = st.text_input("🏢 Account ID", placeholder="Ej: 9695")
    with c3: country = st.selectbox("🌎 País", COUNTRY_OPTIONS)
    def tpl():
        wb=openpyxl.Workbook(); ws=wb.active; ws.title="Tipos de Documento"
        ws.append(["name","entity_type"]); ws.append(["DNI","driver"]); ws.append(["RUC","provider"])
        ws.column_dimensions["A"].width=25; ws.column_dimensions["B"].width=25
        buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf
    st.download_button("📥 Plantilla", data=tpl(), file_name="plantilla_tipos_documento.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    doc_file = st.file_uploader("📂 Excel", type=["xlsx"], key="upload_doctype")
    if doc_file:
        try:
            rows = read_excel_column(doc_file, ["name","entity_type"])
            st.success(f"✅ {len(rows)} tipo(s)")
            with st.expander("Ver"): [st.markdown(f"- **{r['name']}** → `{r['entity_type']}`") for r in rows]
        except Exception as e:
            rows=[]; st.error(str(e))
        if rows and token and account_id and st.button("🚀 Crear", type="primary", key="btn_doctype"):
            results=[]; prog=st.progress(0); status=st.empty()
            for i, row in enumerate(rows):
                status.info(f"Creando: **{row['name']}** ({i+1}/{len(rows)})")
                try:
                    r=requests.post("https://api.simpliroute.com/tms/api/v1/document-types/",
                                    headers={"Authorization":f"Token {token}","Related-Account":account_id,"Content-Type":"application/json"},
                                    json={"name":row["name"],"country":country,"entity_type":row["entity_type"],"status":"active"}, timeout=30)
                    code,resp=r.status_code,r.json()
                except Exception as e: code,resp=None,str(e)
                results.append({"name":row["name"],"code":code,"resp":resp}); prog.progress((i+1)/len(rows))
            status.empty(); prog.empty(); show_results(results, "name")


# ── TMS: TRANSPORTISTAS ───────────────────────────────────────────────────────
def page_tms_transportistas():
    st.title("🚚 Transportistas TMS")
    c1,c2 = st.columns(2)
    with c1: token = st.text_input("🔑 Token", type="password", key="token_transp")
    with c2: account_id = st.text_input("🏢 Account ID", placeholder="Ej: 82761")
    def tpl():
        wb=openpyxl.Workbook(); ws=wb.active; ws.title="Transportistas"
        ws.append(["trade_name","legal_name","tax_id_number"])
        ws.append(["Transportes Sur S.A.","Transportes Sur Sociedad Anónima","20123456789"])
        ws.column_dimensions["A"].width=30; ws.column_dimensions["B"].width=35; ws.column_dimensions["C"].width=20
        buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf
    st.download_button("📥 Plantilla", data=tpl(), file_name="plantilla_transportistas.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    tf = st.file_uploader("📂 Excel", type=["xlsx"], key="upload_transp")
    if tf:
        try:
            rows = read_excel_column(tf, ["trade_name","legal_name","tax_id_number"])
            st.success(f"✅ {len(rows)} transportista(s)")
            with st.expander("Ver"): [st.markdown(f"- **{r['trade_name']}** | {r['legal_name']} | `{r['tax_id_number']}`") for r in rows]
        except Exception as e:
            rows=[]; st.error(str(e))
        if rows and token and account_id and st.button("🚀 Crear", type="primary", key="btn_transp"):
            results=[]; prog=st.progress(0); status=st.empty()
            for i, row in enumerate(rows):
                status.info(f"Creando: **{row['trade_name']}** ({i+1}/{len(rows)})")
                try:
                    r=requests.post("https://api.simpliroute.com/tms/api/v1/transportation-providers/",
                                    headers={"Authorization":f"Token {token}","Content-Type":"application/json"},
                                    json={"trade_name":row["trade_name"],"legal_name":row["legal_name"],
                                          "tax_id_number":row["tax_id_number"],"account_id":int(account_id)}, timeout=30)
                    code,resp=r.status_code,r.json()
                except Exception as e: code,resp=None,str(e)
                results.append({"trade_name":row["trade_name"],"code":code,"resp":resp}); prog.progress((i+1)/len(rows))
            status.empty(); prog.empty(); show_results(results, "trade_name")


# ── ROUTER ────────────────────────────────────────────────────────────────────
if selected == "🧑‍💼 Agregar Seller a Visitas":     page_agregar_seller()
elif selected == "🚛 Flotas":                        page_flotas()
elif selected == "🗺️ Zonas":                         page_zonas()
elif selected == "🔍 Permisos de Usuario":           page_permisos_usuario()
elif selected == "👤 Cambiar Rol de Usuario":         page_cambiar_rol()
elif selected == "⚙️ Configurar Addons":             page_configurar_addons()
elif selected == "🔔 Crear Webhook":                 page_crear_webhook()
elif selected == "🔓 Desbloqueo de Contraseña":      page_desbloqueo()
elif selected == "✏️ Edición de Visitas":            page_edicion_visitas()
elif selected == "🗑️ Eliminación Masiva de Visitas": page_eliminacion_visitas()
elif selected == "🚦 Iniciar / Cerrar Rutas":        page_iniciar_cerrar_rutas()
elif selected == "📍 Análisis de Recorrido GPS":     page_analisis_gps()
elif selected == "🔁 Reenviar Webhooks":             page_reenviar_webhooks()
elif selected == "🏷️ Tipos de Visita y Skills":      page_visit_types_skills()
elif selected == "📡 Validación de GPS":             page_validacion_gps()
elif selected == "📄 Tipos de Documento":            page_tms_document_types()
elif selected == "🚚 Transportistas":                page_tms_transportistas()