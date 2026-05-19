import streamlit as st
import openpyxl
import io
import time
from datetime import date

from utils.helpers import read_excel_column, show_results, mask_token
from utils.api import sr_get, sr_post, sr_patch


def _eta_status(status, i, total, start_time, label):
    elapsed = time.time() - start_time
    rate    = (i + 1) / elapsed if elapsed > 0 else 1
    eta     = (total - (i + 1)) / rate
    status.info(f"{label} ({i+1}/{total}) — ETA: {int(eta)}s")


def page_agregar_seller():
    st.title("🧑‍💼 Agregar Seller a Visitas")
    token = st.text_input("🔑 Token de SimpliRoute", type="password", key="token_seller")

    if st.button("🔍 Consultar Sellers", type="primary", disabled=not token):
        with st.spinner("Consultando..."):
            code, resp = sr_get("/v1/sellers/", token, timeout=60)
        if code == 200:
            if not resp:
                st.warning("⚠️ Esta cuenta no tiene sellers configurados.")
                st.session_state.pop("sellers", None)
            else:
                st.session_state["sellers"] = resp
                st.session_state["seller_token"] = token
                st.success(f"✅ {len(resp)} seller(s)")
        else:
            st.error(f"❌ Error {code}")
            st.session_state.pop("sellers", None)

    if "sellers" in st.session_state:
        sellers = st.session_state["sellers"]
        st.divider()

        def slabel(s):
            name = s.get("name") or s.get("username") or s.get("email") or str(s.get("uuid", ""))
            email = s.get("email", "")
            return f"{name} — {email}" if email and email != name else name

        opciones = {slabel(s): s for s in sellers}
        selected_s = opciones[st.selectbox("Seller:", list(opciones.keys()))]
        seller_uuid = selected_s.get("uuid") or selected_s.get("id")
        st.divider()
        st.caption("Pega los IDs de visita uno por línea:\n```\n799841373\n808472905\n```")
        visit_ids_raw = st.text_area("IDs", placeholder="799841373\n808472905", height=200, label_visibility="collapsed")

        if st.button("💾 Asignar Seller", type="primary", disabled=not visit_ids_raw):
            ids, errs = [], []
            for l in visit_ids_raw.strip().splitlines():
                l = l.strip()
                try:
                    ids.append(int(l))
                except:
                    errs.append(l)
            if errs:
                st.error(f"❌ IDs inválidos: {', '.join(errs)}")
                return
            with st.spinner(f"Asignando a {len(ids)} visitas..."):
                code, _ = sr_patch("/v1/routes/visits/", st.session_state["seller_token"],
                                   [{"id": vid, "seller": seller_uuid} for vid in ids], timeout=120)
            if code in [200, 201]:
                st.success(f"✅ Seller asignado a {len(ids)} visitas")
            else:
                st.error(f"❌ Error {code}")


def page_edicion_visitas():
    st.title("✏️ Edición de Visitas")
    st.markdown("Edita la fecha planificada o la ruta asignada de un lote de visitas.")
    token = st.text_input("🔑 Token de SimpliRoute", type="password", key="token_edit_visits")
    st.divider()
    st.subheader("📋 IDs de visita")
    st.caption("Pega los IDs uno por línea:\n```\n833673298\n837739792\n```")
    visit_ids_raw = st.text_area("IDs de visita", placeholder="833673298\n837739792", height=180, label_visibility="collapsed")
    st.divider()
    st.subheader("📅 Fecha planificada")
    date_action = st.radio("Acción sobre la fecha:", ["No cambiar", "Asignar nueva fecha", "Eliminar fecha"], horizontal=True, key="date_action")
    new_date = None
    if date_action == "Asignar nueva fecha":
        new_date = st.date_input("Selecciona la fecha:", value=date.today(), key="edit_date")
        st.caption(f"Se asignará: **{new_date.strftime('%Y-%m-%d')}**")
    elif date_action == "Eliminar fecha":
        st.warning("⚠️ Se eliminará la fecha planificada de todas las visitas.")
    st.divider()
    st.subheader("🛣️ Ruta")
    route_action = st.radio("Acción sobre la ruta:", ["No cambiar", "Asignar ruta", "Eliminar ruta"], horizontal=True, key="route_action")
    new_route = None
    if route_action == "Asignar ruta":
        new_route = st.text_input("ID de la ruta:", placeholder="5859a5d1-03c5-4152-bcea-500bab2ad47d")
    elif route_action == "Eliminar ruta":
        st.warning("⚠️ Se eliminará la ruta asignada de todas las visitas.")
    st.divider()

    no_changes = date_action == "No cambiar" and route_action == "No cambiar"
    if st.button("🚀 Editar visitas", type="primary", disabled=not (token and visit_ids_raw) or no_changes):
        visit_ids = []
        for l in visit_ids_raw.strip().splitlines():
            l = l.strip()
            if not l: continue
            try: visit_ids.append(int(l))
            except: st.warning(f"⚠️ ID inválido ignorado: {l}")
        if not visit_ids:
            st.error("❌ No se encontraron IDs válidos.")
            return

        def build_payload(vid):
            p = {"id": vid}
            if date_action == "Asignar nueva fecha": p["planned_date"] = new_date.strftime("%Y-%m-%d")
            elif date_action == "Eliminar fecha":    p["planned_date"] = None
            if route_action == "Asignar ruta":       p["route"] = new_route.strip()
            elif route_action == "Eliminar ruta":    p["route"] = None
            return p

        prog = st.progress(0); status = st.empty()
        ok_count = 0; err_count = 0
        batches    = [visit_ids[i:i+500] for i in range(0, len(visit_ids), 500)]
        start_time = time.time()
        for i, batch in enumerate(batches):
            elapsed = time.time() - start_time
            rate    = (i + 1) / elapsed if elapsed > 0 else 1
            eta     = (len(batches) - (i + 1)) / rate
            status.info(f"Editando lote {i+1}/{len(batches)} — {len(batch)} visitas... ({ok_count} editadas) — ETA: {int(eta)}s")
            code, _ = sr_patch("/v1/routes/visits/", token, [build_payload(vid) for vid in batch], timeout=300)
            if code in [200, 201]:
                ok_count += len(batch)
                st.success(f"✅ Lote {i+1}/{len(batches)} — {len(batch)} visitas editadas")
            elif code == 401:
                st.error("❌ Token inválido."); status.empty(); prog.empty(); return
            else:
                st.error(f"❌ Lote {i+1} — Error {code}"); err_count += 1
            prog.progress((i+1)/len(batches))
        status.empty(); prog.empty()
        st.divider()
        if err_count == 0: st.success(f"✅ Completado — **{ok_count} visitas editadas**")
        else: st.warning(f"⚠️ Completado con errores — **{ok_count} editadas**, {err_count} lote(s) fallido(s)")


def page_eliminacion_visitas():
    st.title("🗑️ Eliminación Masiva de Visitas")
    st.error("⚠️ **ADVERTENCIA:** Esta operación elimina visitas directamente desde la base de datos. La eliminación es permanente y no se puede deshacer.")
    token = st.text_input("🔑 Token de SimpliRoute", type="password", key="token_delete")

    def make_template():
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Visitas"
        ws.append(["id"]); ws.append([838112279]); ws.append([838112568])
        ws.column_dimensions["A"].width = 20
        buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf

    st.download_button("📥 Descargar plantilla", data=make_template(),
                       file_name="plantilla_eliminacion_visitas.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    visits_file = st.file_uploader("📂 Sube tu Excel con IDs de visita", type=["xlsx"], key="upload_delete")

    if visits_file:
        try:
            wb = openpyxl.load_workbook(visits_file); ws = wb.active
            headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
            if "id" not in headers: st.error("❌ El Excel debe tener columna 'id'."); return
            idx = headers.index("id")
            visit_ids = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                val = row[idx] if idx < len(row) else None
                if val is not None:
                    try: visit_ids.append(int(val))
                    except: pass
            if not visit_ids: st.error("❌ No se encontraron IDs."); return
            st.info(f"📋 **{len(visit_ids)} visitas** cargadas. Se procesarán en lotes de 2,000.")
        except Exception as e:
            st.error(f"❌ Error: {e}"); return

        confirm = st.checkbox("✅ Confirmo que quiero eliminar estas visitas de forma permanente")
        if confirm and st.button("🗑️ Eliminar visitas", type="primary", disabled=not token):
            batches    = [visit_ids[i:i+2000] for i in range(0, len(visit_ids), 2000)]
            prog = st.progress(0); status = st.empty()
            total_deleted = 0; errors = 0
            start_time = time.time()
            for i, batch in enumerate(batches):
                elapsed = time.time() - start_time
                rate    = (i + 1) / elapsed if elapsed > 0 else 1
                eta     = (len(batches) - (i + 1)) / rate
                status.info(f"Procesando lote {i+1}/{len(batches)} — {len(batch)} visitas... ({total_deleted} eliminadas) — ETA: {int(eta)}s")
                code, _ = sr_post("/v1/bulk/delete/visits/", token, {"visits": batch}, timeout=300)
                if code in [200, 201, 204]:
                    total_deleted += len(batch)
                    st.success(f"✅ Lote {i+1}/{len(batches)} — {len(batch)} visitas eliminadas")
                elif code == 401:
                    st.error("❌ Token inválido."); status.empty(); prog.empty(); return
                else:
                    st.error(f"❌ Lote {i+1} — Error {code}"); errors += 1
                prog.progress((i+1)/len(batches))
            status.empty(); prog.empty()
            st.divider()
            if errors == 0: st.success(f"✅ Completado — **{total_deleted} visitas eliminadas**")
            else: st.warning(f"⚠️ Completado con errores — **{total_deleted} eliminadas**, {errors} lote(s) fallido(s)")


def page_visit_types_skills():
    st.title("🏷️ Tipos de Visita y Skills")
    tab1, tab2, tab3, tab4 = st.tabs([
        "📋 Tipos de Visita",
        "📋 Tipos de Visita Masivo",
        "🔧 Skills",
        "🔧 Skills Masivo",
    ])

    with tab1:
        token_vt = st.text_input("🔑 Token de SimpliRoute", type="password", key="token_vt_single")

        def tpl_vt_single():
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Tipos de Visita"
            ws.append(["label", "key"])
            ws.append(["Entrega express", "entrega_express"])
            ws.append(["Retiro", "retiro"])
            ws.column_dimensions["A"].width = 30
            ws.column_dimensions["B"].width = 30
            buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf

        st.download_button("📥 Plantilla Tipos de Visita", data=tpl_vt_single(),
                           file_name="plantilla_tipos_visita.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key="dl_tpl_vt_single")
        vf_single = st.file_uploader("📂 Excel Tipos de Visita", type=["xlsx"], key="upload_vt_single")
        if vf_single:
            try:
                rows_vt = read_excel_column(vf_single, ["label", "key"])
                st.success(f"✅ {len(rows_vt)} tipo(s) cargados")
                with st.expander("Ver"):
                    for r in rows_vt: st.markdown(f"- **{r['label']}** → `{r['key']}`")
            except Exception as e:
                rows_vt = []; st.error(str(e))
            if rows_vt and token_vt and st.button("🚀 Crear Tipos de Visita", type="primary", key="btn_vt_single"):
                results = []; prog = st.progress(0); status = st.empty()
                start_time = time.time()
                for i, row in enumerate(rows_vt):
                    elapsed = time.time() - start_time
                    rate    = (i + 1) / elapsed if elapsed > 0 else 1
                    eta     = (len(rows_vt) - (i + 1)) / rate
                    status.info(f"Creando: **{row['label']}** ({i+1}/{len(rows_vt)}) — ETA: {int(eta)}s")
                    code, resp = sr_post("/v1/accounts/visit-types/", token_vt,
                                        [{"label": row["label"], "key": row["key"]}])
                    results.append({"label": row["label"], "code": code, "resp": resp})
                    prog.progress((i+1) / len(rows_vt))
                status.empty(); prog.empty()
                show_results(results, "label")

    with tab2:
        st.markdown("Crea tipos de visita en **múltiples cuentas**. Cada fila aplica al token indicado.")

        def tpl_vt_masivo():
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Tipos de Visita Masivo"
            ws.append(["token", "label", "key"])
            ws.append(["TOKEN_CUENTA_1", "Entrega express", "entrega_express"])
            ws.append(["TOKEN_CUENTA_1", "Retiro", "retiro"])
            ws.append(["TOKEN_CUENTA_2", "Entrega programada", "entrega_programada"])
            ws.column_dimensions["A"].width = 40
            ws.column_dimensions["B"].width = 30
            ws.column_dimensions["C"].width = 30
            buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf

        st.download_button("📥 Plantilla Masiva", data=tpl_vt_masivo(),
                           file_name="plantilla_tipos_visita_masivo.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key="dl_tpl_vt_masivo")
        vf_masivo = st.file_uploader("📂 Excel Masivo", type=["xlsx"], key="upload_vt_masivo")
        if vf_masivo:
            try:
                rows_vt_masivo = read_excel_column(vf_masivo, ["token", "label", "key"])
                cuentas_vt = len(set(r["token"] for r in rows_vt_masivo if r["token"]))
                st.success(f"✅ {len(rows_vt_masivo)} tipo(s) para {cuentas_vt} cuenta(s)")
                with st.expander("Ver"):
                    for r in rows_vt_masivo:
                        st.markdown(f"- `{mask_token(r['token'])}` → **{r['label']}** / `{r['key']}`")
            except Exception as e:
                rows_vt_masivo = []; st.error(str(e))
            if rows_vt_masivo and st.button("🚀 Crear Masivo", type="primary", key="btn_vt_masivo"):
                results = []; prog = st.progress(0); status = st.empty()
                start_time = time.time()
                for i, row in enumerate(rows_vt_masivo):
                    tok_p   = mask_token(row["token"])
                    elapsed = time.time() - start_time
                    rate    = (i + 1) / elapsed if elapsed > 0 else 1
                    eta     = (len(rows_vt_masivo) - (i + 1)) / rate
                    status.info(f"Creando: **{row['label']}** en `{tok_p}` ({i+1}/{len(rows_vt_masivo)}) — ETA: {int(eta)}s")
                    code, resp = sr_post("/v1/accounts/visit-types/", row["token"],
                                        [{"label": row["label"], "key": row["key"]}])
                    results.append({"label": f"{row['label']} ({tok_p})", "code": code, "resp": resp})
                    prog.progress((i+1) / len(rows_vt_masivo))
                status.empty(); prog.empty()
                show_results(results, "label")

    with tab3:
        token_sk = st.text_input("🔑 Token de SimpliRoute", type="password", key="token_sk_single")

        def tpl_skills_single():
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Skills"
            ws.append(["skill"])
            ws.append(["Manejo de carga pesada"])
            ws.append(["Refrigeración"])
            ws.column_dimensions["A"].width = 35
            buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf

        st.download_button("📥 Plantilla Skills", data=tpl_skills_single(),
                           file_name="plantilla_skills.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key="dl_tpl_sk_single")
        sf_single = st.file_uploader("📂 Excel Skills", type=["xlsx"], key="upload_sk_single")
        if sf_single:
            try:
                rows_sk = read_excel_column(sf_single, ["skill"])
                st.success(f"✅ {len(rows_sk)} skill(s) cargados")
                with st.expander("Ver"):
                    for r in rows_sk: st.markdown(f"- {r['skill']}")
            except Exception as e:
                rows_sk = []; st.error(str(e))
            if rows_sk and token_sk and st.button("🚀 Crear Skills", type="primary", key="btn_sk_single"):
                results = []; prog = st.progress(0); status = st.empty()
                start_time = time.time()
                for i, row in enumerate(rows_sk):
                    elapsed = time.time() - start_time
                    rate    = (i + 1) / elapsed if elapsed > 0 else 1
                    eta     = (len(rows_sk) - (i + 1)) / rate
                    status.info(f"Creando: **{row['skill']}** ({i+1}/{len(rows_sk)}) — ETA: {int(eta)}s")
                    code, resp = sr_post("/v1/routes/skills/", token_sk, {"skill": row["skill"]})
                    results.append({"skill": row["skill"], "code": code, "resp": resp})
                    prog.progress((i+1) / len(rows_sk))
                status.empty(); prog.empty()
                show_results(results, "skill")

    with tab4:
        st.markdown("Crea skills en **múltiples cuentas**. Cada fila aplica al token indicado.")

        def tpl_skills_masivo():
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Skills Masivo"
            ws.append(["token", "skill"])
            ws.append(["TOKEN_CUENTA_1", "Manejo de carga pesada"])
            ws.append(["TOKEN_CUENTA_1", "Refrigeración"])
            ws.append(["TOKEN_CUENTA_2", "Conducción nocturna"])
            ws.column_dimensions["A"].width = 40
            ws.column_dimensions["B"].width = 35
            buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf

        st.download_button("📥 Plantilla Masiva", data=tpl_skills_masivo(),
                           file_name="plantilla_skills_masivo.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key="dl_tpl_sk_masivo")
        sf_masivo = st.file_uploader("📂 Excel Masivo", type=["xlsx"], key="upload_sk_masivo")
        if sf_masivo:
            try:
                rows_sk_masivo = read_excel_column(sf_masivo, ["token", "skill"])
                cuentas_sk = len(set(r["token"] for r in rows_sk_masivo if r["token"]))
                st.success(f"✅ {len(rows_sk_masivo)} skill(s) para {cuentas_sk} cuenta(s)")
                with st.expander("Ver"):
                    for r in rows_sk_masivo:
                        st.markdown(f"- `{mask_token(r['token'])}` → **{r['skill']}**")
            except Exception as e:
                rows_sk_masivo = []; st.error(str(e))
            if rows_sk_masivo and st.button("🚀 Crear Masivo", type="primary", key="btn_sk_masivo"):
                results = []; prog = st.progress(0); status = st.empty()
                start_time = time.time()
                for i, row in enumerate(rows_sk_masivo):
                    tok_p   = mask_token(row["token"])
                    elapsed = time.time() - start_time
                    rate    = (i + 1) / elapsed if elapsed > 0 else 1
                    eta     = (len(rows_sk_masivo) - (i + 1)) / rate
                    status.info(f"Creando: **{row['skill']}** en `{tok_p}` ({i+1}/{len(rows_sk_masivo)}) — ETA: {int(eta)}s")
                    code, resp = sr_post("/v1/routes/skills/", row["token"], {"skill": row["skill"]})
                    results.append({"skill": f"{row['skill']} ({tok_p})", "code": code, "resp": resp})
                    prog.progress((i+1) / len(rows_sk_masivo))
                status.empty(); prog.empty()
                show_results(results, "skill")
