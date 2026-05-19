import streamlit as st
import openpyxl
import json
import io
from datetime import date, datetime

from utils.helpers import haversine_m
from utils.api import sr_get, API_GW


def page_analisis_gps():
    st.title("📍 Análisis de Recorrido GPS")
    st.markdown("Carga un JSON de puntos GPS para analizar el recorrido, detectar anomalías y comparar contra un punto fijo.")
    json_file = st.file_uploader("📂 Sube tu archivo JSON", type=["json"], key="upload_gps_analysis")
    if not json_file: return
    try:
        data = json.loads(json_file.read())
        if not isinstance(data, list) or len(data) < 2:
            st.error("❌ El JSON debe ser una lista con al menos 2 puntos."); return
    except Exception as e:
        st.error(f"❌ Error leyendo el JSON: {e}"); return

    st.success(f"✅ {len(data)} puntos cargados")
    st.divider()
    st.subheader("🛣️ Análisis de recorrido")
    st.caption("Puntos anómalos: velocidad implícita entre puntos consecutivos **> 500 km/h**.")

    VELOCIDAD_MAX = 500
    total_bruto = 0; total_limpio = 0; anomalos = []; detalles = []

    for i in range(1, len(data)):
        p1, p2 = data[i-1], data[i]
        try:
            lat1, lon1 = float(p1["latitude"]), float(p1["longitude"])
            lat2, lon2 = float(p2["latitude"]), float(p2["longitude"])
            dist_m = haversine_m(lat1, lon1, lat2, lon2)
            velocidad_kmh = None
            try:
                t1 = datetime.fromisoformat(p1["timestamp"].replace("Z", "+00:00"))
                t2 = datetime.fromisoformat(p2["timestamp"].replace("Z", "+00:00"))
                seg = abs((t2-t1).total_seconds())
                if seg > 0: velocidad_kmh = (dist_m/1000) / (seg/3600)
            except: pass
            es_anomalo = velocidad_kmh is not None and velocidad_kmh > VELOCIDAD_MAX
            total_bruto += dist_m
            if not es_anomalo: total_limpio += dist_m
            detalles.append({"index": i, "timestamp": p2.get("timestamp", "—"), "lat": lat2, "lon": lon2,
                             "dist_m": round(dist_m, 2),
                             "velocidad_kmh": round(velocidad_kmh, 1) if velocidad_kmh else None,
                             "anomalo": es_anomalo})
            if es_anomalo:
                anomalos.append({"index": i, "timestamp": p2.get("timestamp", "—"), "lat": lat2, "lon": lon2,
                                 "dist_km": round(dist_m/1000, 2), "velocidad_kmh": round(velocidad_kmh, 1)})
        except: continue

    col1, col2, col3 = st.columns(3)
    col1.metric("📏 Total bruto",    f"{total_bruto/1000:.2f} km")
    col2.metric("✅ Total limpio",   f"{total_limpio/1000:.2f} km")
    col3.metric("⚠️ Puntos anómalos", len(anomalos))

    if anomalos:
        top10 = sorted(anomalos, key=lambda x: x["dist_km"], reverse=True)[:10]
        st.markdown(f"**{len(anomalos)} puntos anómalos — Top 10:**")
        for a in top10:
            st.error(f"⚠️ Índice {a['index']} | `{a['timestamp']}` | `{a['lat']}, {a['lon']}` | {a['dist_km']} km | {a['velocidad_kmh']} km/h")

        def make_excel_anomalos(items):
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Anomalos"
            ws.append(["index", "timestamp", "latitude", "longitude", "distancia_km", "velocidad_kmh"])
            for a in sorted(items, key=lambda x: x["dist_km"], reverse=True):
                ws.append([a["index"], a["timestamp"], a["lat"], a["lon"], a["dist_km"], a["velocidad_kmh"]])
            ws.column_dimensions["B"].width = 25
            buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf

        st.download_button("⬇️ Descargar Excel anomalos", data=make_excel_anomalos(anomalos),
                           file_name="puntos_anomalos.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    else:
        st.success("✅ No se detectaron puntos anómalos.")

    def make_excel_analisis(detalles):
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Recorrido"
        ws.append(["index", "timestamp", "latitude", "longitude", "distancia_m", "velocidad_kmh", "anomalo"])
        for d in detalles:
            ws.append([d["index"], d["timestamp"], d["lat"], d["lon"], d["dist_m"],
                       d["velocidad_kmh"], "Sí" if d["anomalo"] else "No"])
        ws.column_dimensions["B"].width = 25
        buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf

    st.download_button("⬇️ Descargar Excel recorrido", data=make_excel_analisis(detalles),
                       file_name="analisis_recorrido.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    st.divider()
    st.subheader("📌 Comparación con punto fijo")
    col_a, col_b, col_c = st.columns(3)
    with col_a: ref_lat = st.text_input("🌐 Latitud",  placeholder="-22.797010")
    with col_b: ref_lon = st.text_input("🌐 Longitud", placeholder="-43.323240")
    with col_c: radio_m = st.number_input("📏 Radio (metros)", min_value=1, value=500, step=50)

    if st.button("🔍 Comparar puntos", type="primary", disabled=not (ref_lat and ref_lon)):
        try: rlat, rlon = float(ref_lat), float(ref_lon)
        except: st.error("❌ Coordenadas inválidas."); return
        dentro = []; fuera = 0
        for i, p in enumerate(data):
            try:
                plat, plon = float(p["latitude"]), float(p["longitude"])
                dist = haversine_m(rlat, rlon, plat, plon)
                if dist <= radio_m:
                    dentro.append({"index": i, "timestamp": p.get("timestamp", "—"),
                                   "lat": plat, "lon": plon, "dist_m": round(dist, 1)})
                else: fuera += 1
            except: continue
        col_x, col_y = st.columns(2)
        col_x.metric(f"✅ Dentro ({radio_m}m)", len(dentro))
        col_y.metric("❌ Fuera", fuera)
        if dentro:
            with st.expander(f"Ver {len(dentro)} puntos"):
                for p in dentro[:50]:
                    st.markdown(f"- Índice `{p['index']}` | `{p['timestamp']}` | `{p['lat']}, {p['lon']}` | **{p['dist_m']} m**")
                if len(dentro) > 50: st.caption(f"... y {len(dentro)-50} más.")

            def make_excel_radio(puntos):
                wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Puntos cercanos"
                ws.append(["index", "timestamp", "latitude", "longitude", "distancia_al_punto_m"])
                for p in puntos:
                    ws.append([p["index"], p["timestamp"], p["lat"], p["lon"], p["dist_m"]])
                ws.column_dimensions["B"].width = 25
                buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf

            st.download_button("⬇️ Descargar Excel", data=make_excel_radio(dentro),
                               file_name=f"puntos_dentro_{radio_m}m.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        else:
            st.info(f"ℹ️ Ningún punto dentro de {radio_m} metros.")


def page_validacion_gps():
    st.title("📡 Validación de GPS")
    st.markdown("Consulta si un vehículo o conductor tiene registros de ubicación para una fecha determinada.")
    col1, col2 = st.columns(2)
    with col1: token = st.text_input("🔑 Token de SimpliRoute", type="password", key="token_gps")
    with col2: selected_date = st.date_input("📅 Fecha", value=date.today(), key="gps_date")
    date_str = selected_date.strftime("%Y-%m-%d")
    modo = st.radio("🔍 Modo de consulta", ["Por tipo (Vehículo o Conductor)", "Por par (Driver + Vehículo)"], horizontal=True)
    st.divider()

    if modo == "Por tipo (Vehículo o Conductor)":
        entity_type = st.selectbox("Consultar por", ["Vehículo", "Conductor"])
        st.caption("Pega los IDs uno por línea:\n```\n568025\n568026\n```")
        entity_ids_raw = st.text_area("IDs", placeholder="568025\n568026", height=150,
                                      label_visibility="collapsed", key="single_ids")
        if st.button("🔍 Consultar GPS", type="primary", disabled=not (token and entity_ids_raw)):
            param_key  = "vehicle_id" if entity_type == "Vehículo" else "driver_id"
            entity_ids = [l.strip() for l in entity_ids_raw.strip().splitlines() if l.strip()]
            if not entity_ids: st.error("❌ No se encontraron IDs."); return
            con_data = []; sin_data = []; all_records = []
            prog = st.progress(0); status = st.empty()
            for i, eid in enumerate(entity_ids):
                status.info(f"Consultando {entity_type.lower()} **{eid}** ({i+1}/{len(entity_ids)})")
                code, data = sr_get(f"/v1/tracking/locations/{date_str}/?{param_key}={eid}", token, base=API_GW)
                if code == 401: st.error("❌ Token inválido."); status.empty(); prog.empty(); return
                elif code != 200: st.error(f"❌ **{eid}** — Error {code}")
                elif not data: sin_data.append(eid)
                else:
                    con_data.append({"id": eid, "count": len(data)})
                    for rec in data: rec["entity_id"] = eid; all_records.append(rec)
                prog.progress((i+1)/len(entity_ids))
            status.empty(); prog.empty()
            st.divider()
            col_a, col_b = st.columns(2)
            with col_a:
                st.success(f"✅ Con data GPS: **{len(con_data)}**")
                for item in con_data: st.markdown(f"- ID `{item['id']}` — {item['count']} registros")
            with col_b:
                st.error(f"❌ Sin data GPS: **{len(sin_data)}**")
                for eid in sin_data: st.markdown(f"- ID `{eid}`")
            if all_records:
                def make_excel_single(records):
                    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "GPS"
                    ws.append(["entity_id","timestamp","latitude","longitude","activity_type","type","id","accuracy"])
                    for rec in records:
                        ws.append([rec.get("entity_id",""), rec.get("timestamp",""), rec.get("latitude",""),
                                   rec.get("longitude",""), rec.get("activity_type",""), rec.get("type",""),
                                   rec.get("id",""), rec.get("accuracy","")])
                    ws.column_dimensions["B"].width = 25
                    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf
                col_dl1, col_dl2 = st.columns(2)
                with col_dl1:
                    st.download_button("⬇️ Descargar Excel", data=make_excel_single(all_records),
                                       file_name=f"gps_{entity_type.lower()}_{date_str}.xlsx",
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                with col_dl2:
                    st.download_button("⬇️ Descargar JSON",
                                       data=json.dumps(all_records, ensure_ascii=False, indent=2),
                                       file_name=f"gps_{entity_type.lower()}_{date_str}.json",
                                       mime="application/json")
    else:
        col3, col4 = st.columns(2)
        with col3:
            st.caption("ID de conductor:")
            driver_ids_raw = st.text_area("Driver IDs", placeholder="523842", height=80,
                                          label_visibility="collapsed", key="driver_ids")
        with col4:
            st.caption("ID de vehículo:")
            vehicle_ids_raw = st.text_area("Vehicle IDs", placeholder="674797", height=80,
                                           label_visibility="collapsed", key="vehicle_ids")
        if st.button("🔍 Consultar GPS", type="primary",
                     disabled=not (token and driver_ids_raw and vehicle_ids_raw), key="btn_gps_pair"):
            driver_ids  = [l.strip() for l in driver_ids_raw.strip().splitlines()  if l.strip()]
            vehicle_ids = [l.strip() for l in vehicle_ids_raw.strip().splitlines() if l.strip()]
            if len(driver_ids) != len(vehicle_ids):
                st.error(f"❌ Cantidad no coincide: {len(driver_ids)} conductores vs {len(vehicle_ids)} vehículos."); return
            con_data = []; sin_data = []; all_records = []
            prog = st.progress(0); status = st.empty()
            for i, (did, vid) in enumerate(zip(driver_ids, vehicle_ids)):
                status.info(f"Consultando conductor **{did}** / vehículo **{vid}** ({i+1}/{len(driver_ids)})")
                code, data = sr_get(f"/v1/tracking/locations/{date_str}/?driver_id={did}&vehicle_id={vid}",
                                    token, base=API_GW)
                if code == 401: st.error("❌ Token inválido."); status.empty(); prog.empty(); return
                elif code != 200: st.error(f"❌ **{did}/{vid}** — Error {code}")
                elif not data: sin_data.append(f"Driver {did} / Vehículo {vid}")
                else:
                    con_data.append({"driver": did, "vehicle": vid, "count": len(data)})
                    for rec in data: rec["driver_id"] = did; rec["vehicle_id"] = vid; all_records.append(rec)
                prog.progress((i+1)/len(driver_ids))
            status.empty(); prog.empty()
            col_a, col_b = st.columns(2)
            with col_a:
                st.success(f"✅ Con data GPS: **{len(con_data)}**")
                for item in con_data:
                    st.markdown(f"- Driver `{item['driver']}` / Vehículo `{item['vehicle']}` — {item['count']} registros")
            with col_b:
                st.error(f"❌ Sin data GPS: **{len(sin_data)}**")
                for pair in sin_data: st.markdown(f"- {pair}")
            if all_records:
                def make_excel_pair(records):
                    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "GPS"
                    ws.append(["driver_id","vehicle_id","timestamp","latitude","longitude","activity_type","type","id","accuracy"])
                    for rec in records:
                        ws.append([rec.get("driver_id",""), rec.get("vehicle_id",""), rec.get("timestamp",""),
                                   rec.get("latitude",""), rec.get("longitude",""), rec.get("activity_type",""),
                                   rec.get("type",""), rec.get("id",""), rec.get("accuracy","")])
                    ws.column_dimensions["C"].width = 25
                    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf
                col_dl3, col_dl4 = st.columns(2)
                with col_dl3:
                    st.download_button("⬇️ Descargar Excel", data=make_excel_pair(all_records),
                                       file_name=f"gps_pares_{date_str}.xlsx",
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                with col_dl4:
                    st.download_button("⬇️ Descargar JSON",
                                       data=json.dumps(all_records, ensure_ascii=False, indent=2),
                                       file_name=f"gps_pares_{date_str}.json", mime="application/json")
