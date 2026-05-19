import math
import re
import io
import openpyxl


def haversine_m(lat1, lon1, lat2, lon2):
    R = 6371000
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lon2 - lon1)
    a = math.sin(dphi/2)**2 + math.cos(phi1)*math.cos(phi2)*math.sin(dlambda/2)**2
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def decode_file(raw_bytes):
    for enc in ["utf-8", "latin-1", "cp1252"]:
        try:
            return raw_bytes.decode(enc, errors="ignore")
        except Exception:
            continue
    return raw_bytes.decode("utf-8", errors="ignore")


def strip_rtf_codes(text):
    text = re.sub(r"\\'[0-9a-fA-F]{2}", "", text)
    text = re.sub(r"\\[a-zA-Z]+\d*\s?", " ", text)
    text = re.sub(r"[{}]", "", text)
    text = re.sub(r" {2,}", " ", text)
    return text


def parse_polygons(text):
    polygons = []
    for block in re.findall(r'<Placemark>(.*?)</Placemark>', text, re.DOTALL):
        if '<Polygon>' not in block and '<Polygon ' not in block:
            continue
        name_match = re.search(r'<name>\s*(.*?)\s*</name>', block)
        name = name_match.group(1).strip() if name_match else "Sin nombre"
        coords_match = re.search(r'<coordinates>\s*(.*?)\s*</coordinates>', block, re.DOTALL)
        if not coords_match:
            continue
        points = []
        for token_str in coords_match.group(1).strip().split():
            parts = token_str.strip().split(",")
            if len(parts) < 2:
                continue
            try:
                a, b = float(parts[0]), float(parts[1])
            except ValueError:
                continue
            lng, lat = a, b
            if abs(lat) > 90:
                lng, lat = lat, lng
            if -90 <= lat <= 90 and -180 <= lng <= 180:
                points.append((lat, lng))
        if points:
            polygons.append({"name": name, "coords": points})
    return polygons


def coords_to_str(coords):
    return "[" + ",".join(f"{{'lat':'{lat:.6f}','lng':'{lng:.6f}'}}" for lat, lng in coords) + "]"


def generate_excel_zones(polygons):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Hoja 1"
    ws.append(["is_name", "is_coordinates"])
    for p in polygons:
        ws.append([p["name"], coords_to_str(p["coords"])])
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 80
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf


def read_excel_column(file, columns):
    wb = openpyxl.load_workbook(file); ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        entry = {}
        for col in columns:
            idx = headers.index(col) if col in headers else -1
            val = row[idx] if 0 <= idx < len(row) else None
            entry[col] = str(val).strip() if val is not None else ""
        if any(entry[c] for c in columns):
            rows.append(entry)
    return rows


def show_results(results, name_key):
    import streamlit as st
    ok = sum(1 for r in results if r["code"] in [200, 201])
    st.markdown(f"✅ **{ok} creados** | ❌ **{len(results)-ok} con error**")
    for r in results:
        if r["code"] in [200, 201]:
            st.success(f"✅ {r[name_key]} — OK")
        elif r["code"] == 400:
            st.warning(f"⚠️ {r[name_key]} — Ya existe o datos inválidos")
        elif r["code"] == 401:
            st.error(f"❌ {r[name_key]} — Token inválido")
        elif r["code"] is None:
            st.error(f"❌ {r[name_key]} — Sin conexión")
        else:
            st.error(f"❌ {r[name_key]} — Error {r['code']}")


def mask_token(token):
    if len(token) > 10:
        return token[:6] + "..." + token[-4:]
    return token
